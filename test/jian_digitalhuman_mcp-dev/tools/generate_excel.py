#!/usr/bin/env python3
"""
生成Excel文件工具
根据数据库中的数据，按照模板生成Excel文件
"""

import os
import logging
from typing import Optional, Dict, Any
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys
from pathlib import Path

# 添加项目根目录到路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from database import (
    CollectionRecord,
    ContractRecord,
    get_session,
)
from mcp_instance import mcp

logger = logging.getLogger(__name__)

# 模板文件路径
TEMPLATE_PATH = Path(__file__).parent.parent / "template" / "26年2月份预计回款、合同表26.1.28.xlsx"


def format_number(value: Any) -> float:
    """格式化数字，确保返回浮点数"""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def get_collection_data_by_person(session, month: str) -> Dict[str, list]:
    """按负责人分组获取回款数据"""
    records = (
        session.query(CollectionRecord)
        .filter(CollectionRecord.month == month)
        .order_by(
            CollectionRecord.responsible_person,
            CollectionRecord.project_type,
            CollectionRecord.id,
        )
        .all()
    )
    
    grouped = {}
    for record in records:
        person = record.responsible_person
        if person not in grouped:
            grouped[person] = []
        grouped[person].append(record)
    
    return grouped


def get_contract_data_by_person(session, month: str) -> Dict[str, list]:
    """按负责人分组获取合同数据"""
    records = (
        session.query(ContractRecord)
        .filter(ContractRecord.month == month)
        .order_by(
            ContractRecord.responsible_person,
            ContractRecord.project_name,
            ContractRecord.id,
        )
        .all()
    )
    
    grouped = {}
    for record in records:
        person = record.responsible_person
        if person not in grouped:
            grouped[person] = []
        grouped[person].append(record)
    
    return grouped


def write_collection_section(ws, start_row: int, data: Dict[str, list], month: str):
    """写入回款部分数据"""
    current_row = start_row
    
    # 表头（假设从第2行开始，第1行是标题）
    header_row = start_row
    headers = [
        "负责人",
        "项目类型",
        "项目名称",
        "月初预计可能回款",
        "月初预计确定回款",
        "可能回款",
        "确定回款",
        "实际回款",
        "未回款金额",
        "未完成原因",
        "解决办法",
    ]
    
    # 写入表头（如果需要）
    # for col_idx, header in enumerate(headers, start=1):
    #     ws.cell(row=header_row, column=col_idx, value=header)
    
    # 写入数据
    for person, records in data.items():
        # 按项目类型分组
        by_type = {}
        for record in records:
            if record.is_subtotal:
                continue  # 跳过小计行，后面会计算
            project_type = record.project_type
            if project_type not in by_type:
                by_type[project_type] = []
            by_type[project_type].append(record)
        
        # 写入每个负责人的数据
        for project_type, type_records in by_type.items():
            # 写入负责人和项目类型（合并单元格）
            person_start_row = current_row + 1
            
            for record in type_records:
                current_row += 1
                ws.cell(row=current_row, column=1, value=record.responsible_person)
                ws.cell(row=current_row, column=2, value=record.project_type)
                ws.cell(row=current_row, column=3, value=record.project_name)
                ws.cell(row=current_row, column=4, value=format_number(record.estimated_possible_at_start))
                ws.cell(row=current_row, column=5, value=format_number(record.estimated_confirmed_at_start))
                ws.cell(row=current_row, column=6, value=format_number(record.possible_collection))
                ws.cell(row=current_row, column=7, value=format_number(record.confirmed_collection))
                ws.cell(row=current_row, column=8, value=format_number(record.actual_collection))
                ws.cell(row=current_row, column=9, value=format_number(record.uncollected_amount))
                ws.cell(row=current_row, column=10, value=record.reason_for_non_completion or "")
                ws.cell(row=current_row, column=11, value=record.solution or "")
            
            # 合并负责人和项目类型的单元格
            if current_row > person_start_row:
                ws.merge_cells(f"A{person_start_row}:A{current_row}")
                ws.merge_cells(f"B{person_start_row}:B{current_row}")
    
    return current_row


def write_contract_section(ws, start_row: int, data: Dict[str, list], month: str):
    """写入合同部分数据"""
    current_row = start_row
    
    # 写入数据
    for person, records in data.items():
        person_start_row = current_row + 1
        
        for record in records:
            if record.is_subtotal:
                continue  # 跳过小计行
            current_row += 1
            ws.cell(row=current_row, column=1, value=record.responsible_person)
            ws.cell(row=current_row, column=2, value=record.company_name or "")
            ws.cell(row=current_row, column=3, value=record.project_name)
            ws.cell(row=current_row, column=4, value=format_number(record.estimated_possible_at_start))
            ws.cell(row=current_row, column=5, value=format_number(record.estimated_confirmed_at_start))
            ws.cell(row=current_row, column=6, value=format_number(record.possible_contract))
            ws.cell(row=current_row, column=7, value=format_number(record.confirmed_contract))
            ws.cell(row=current_row, column=8, value=format_number(record.actual_contract))
            ws.cell(row=current_row, column=9, value=record.completion_status or "")
        
        # 合并负责人单元格
        if current_row > person_start_row:
            ws.merge_cells(f"A{person_start_row}:A{current_row}")
    
    return current_row


@mcp.tool()
async def generate_financial_excel(
    month: str = None,
    output_path: str = None,
    template_path: str = None,
) -> str:
    """
    根据数据库数据生成Excel文件
    
    Args:
        month: 月份，格式：YYYY-MM（如：2026-02）。如果不提供，使用当前月份
        output_path: 输出文件路径。如果不提供，使用 data/ 目录下自动生成的文件名
        template_path: 模板文件路径。如果不提供，使用默认模板
    
    Returns:
        生成结果信息，包含文件路径
    """
    try:
        # 确定月份
        if not month:
            month = datetime.now().strftime("%Y-%m")
        
        # 确定模板路径
        if not template_path:
            template_path = str(TEMPLATE_PATH)
        
        if not os.path.exists(template_path):
            return json.dumps(
                {
                    "success": False,
                    "message": f"模板文件不存在: {template_path}",
                },
                ensure_ascii=False,
            )
        
        # 确定输出路径
        if not output_path:
            year_month = month.replace("-", "年") + "月"
            output_filename = f"{year_month}份预计回款、合同表.xlsx"
            output_dir = Path(__file__).parent.parent / "data"
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = str(output_dir / output_filename)
        
        # 加载模板
        wb = load_workbook(template_path)
        
        # 获取第一个工作表（或指定工作表）
        ws = wb.active
        
        # 获取数据库数据
        session = get_session()
        try:
            collection_data = get_collection_data_by_person(session, month)
            contract_data = get_contract_data_by_person(session, month)
        finally:
            session.close()
        
        # 查找数据区域的起始行
        # 假设回款数据从第3行开始（第1行是标题，第2行是表头）
        collection_start_row = 3
        
        # 查找合同部分的起始行（需要先找到回款部分的结束位置）
        # 这里简化处理，假设合同部分在回款部分之后
        # 实际应该根据模板结构来确定
        
        # 写入回款数据
        if collection_data:
            # 清空原有数据区域（保留表头）
            # 这里需要根据实际模板结构来调整
            write_collection_section(ws, collection_start_row, collection_data, month)
        
        # 写入合同数据
        if contract_data:
            # 需要找到合同部分的起始行
            # 这里简化处理，假设在回款数据之后
            contract_start_row = collection_start_row + 50  # 临时值，需要根据实际情况调整
            write_contract_section(ws, contract_start_row, contract_data, month)
        
        # 保存文件
        wb.save(output_path)
        
        import json
        return json.dumps(
            {
                "success": True,
                "message": f"成功生成Excel文件",
                "file_path": output_path,
                "month": month,
            },
            ensure_ascii=False,
        )
    
    except Exception as e:
        logger.error(f"生成Excel文件失败: {e}", exc_info=True)
        import json
        return json.dumps(
            {"success": False, "message": f"生成失败: {str(e)}"},
            ensure_ascii=False,
        )


@mcp.tool()
async def generate_excel_from_template(
    month: str = None,
    output_path: str = None,
) -> str:
    """
    从模板生成Excel文件（简化版本，直接复制模板并填充数据）
    
    Args:
        month: 月份，格式：YYYY-MM（如：2026-02）
        output_path: 输出文件路径
    
    Returns:
        生成结果信息
    """
    import json
    import shutil
    
    try:
        # 确定月份
        if not month:
            month = datetime.now().strftime("%Y-%m")
        
        # 确定输出路径
        if not output_path:
            year_month = month.replace("-", "年") + "月"
            output_filename = f"{year_month}份预计回款、合同表.xlsx"
            output_dir = Path(__file__).parent.parent / "data"
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = str(output_dir / output_filename)
        
        # 复制模板
        template_path = str(TEMPLATE_PATH)
        if not os.path.exists(template_path):
            return json.dumps(
                {
                    "success": False,
                    "message": f"模板文件不存在: {template_path}",
                },
                ensure_ascii=False,
            )
        
        shutil.copy2(template_path, output_path)
        
        # 加载工作簿
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 获取数据库数据
        session = get_session()
        try:
            # 获取回款数据
            collection_records = (
                session.query(CollectionRecord)
                .filter(CollectionRecord.month == month)
                .order_by(
                    CollectionRecord.responsible_person,
                    CollectionRecord.project_type,
                    CollectionRecord.id,
                )
                .all()
            )
            
            # 获取合同数据
            contract_records = (
                session.query(ContractRecord)
                .filter(ContractRecord.month == month)
                .order_by(
                    ContractRecord.responsible_person,
                    ContractRecord.project_name,
                    ContractRecord.id,
                )
                .all()
            )
        finally:
            session.close()
        
        # 写入回款数据（从第3行开始，假设第1行是标题，第2行是表头）
        row = 3
        current_person = None
        current_type = None
        person_start_row = None
        type_start_row = None
        
        for record in collection_records:
            if record.is_subtotal:
                continue
            
            # 检查是否需要合并单元格
            if current_person != record.responsible_person:
                if person_start_row and row > person_start_row:
                    ws.merge_cells(f"A{person_start_row}:A{row-1}")
                current_person = record.responsible_person
                person_start_row = row
            
            if current_type != record.project_type:
                if type_start_row and row > type_start_row:
                    ws.merge_cells(f"B{type_start_row}:B{row-1}")
                current_type = record.project_type
                type_start_row = row
            
            # 写入数据
            ws.cell(row=row, column=1, value=record.responsible_person)
            ws.cell(row=row, column=2, value=record.project_type)
            ws.cell(row=row, column=3, value=record.project_name)
            ws.cell(row=row, column=4, value=format_number(record.estimated_possible_at_start))
            ws.cell(row=row, column=5, value=format_number(record.estimated_confirmed_at_start))
            ws.cell(row=row, column=6, value=format_number(record.possible_collection))
            ws.cell(row=row, column=7, value=format_number(record.confirmed_collection))
            ws.cell(row=row, column=8, value=format_number(record.actual_collection))
            ws.cell(row=row, column=9, value=format_number(record.uncollected_amount))
            ws.cell(row=row, column=10, value=record.reason_for_non_completion or "")
            ws.cell(row=row, column=11, value=record.solution or "")
            
            row += 1
        
        # 合并最后一批单元格
        if person_start_row and row > person_start_row:
            ws.merge_cells(f"A{person_start_row}:A{row-1}")
        if type_start_row and row > type_start_row:
            ws.merge_cells(f"B{type_start_row}:B{row-1}")
        
        # 查找合同部分的起始行（假设在回款部分之后，需要根据实际模板调整）
        # 这里简化处理，假设合同部分从第50行开始
        contract_start_row = 50
        row = contract_start_row
        
        current_person = None
        person_start_row = None
        
        for record in contract_records:
            if record.is_subtotal:
                continue
            
            if current_person != record.responsible_person:
                if person_start_row and row > person_start_row:
                    ws.merge_cells(f"A{person_start_row}:A{row-1}")
                current_person = record.responsible_person
                person_start_row = row
            
            # 写入合同数据
            ws.cell(row=row, column=1, value=record.responsible_person)
            ws.cell(row=row, column=2, value=record.company_name or "")
            ws.cell(row=row, column=3, value=record.project_name)
            ws.cell(row=row, column=4, value=format_number(record.estimated_possible_at_start))
            ws.cell(row=row, column=5, value=format_number(record.estimated_confirmed_at_start))
            ws.cell(row=row, column=6, value=format_number(record.possible_contract))
            ws.cell(row=row, column=7, value=format_number(record.confirmed_contract))
            ws.cell(row=row, column=8, value=format_number(record.actual_contract))
            ws.cell(row=row, column=9, value=record.completion_status or "")
            
            row += 1
        
        # 合并最后一批单元格
        if person_start_row and row > person_start_row:
            ws.merge_cells(f"A{person_start_row}:A{row-1}")
        
        # 保存文件
        wb.save(output_path)
        
        return json.dumps(
            {
                "success": True,
                "message": "成功生成Excel文件",
                "file_path": output_path,
                "month": month,
                "collection_count": len(collection_records),
                "contract_count": len(contract_records),
            },
            ensure_ascii=False,
        )
    
    except Exception as e:
        logger.error(f"生成Excel文件失败: {e}", exc_info=True)
        return json.dumps(
            {"success": False, "message": f"生成失败: {str(e)}"},
            ensure_ascii=False,
        )
