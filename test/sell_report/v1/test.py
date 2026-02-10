import json
from pathlib import Path
from openpyxl import Workbook, load_workbook

TABLE_PATH = Path("data/回款表.xlsx")
TEMPLATE_PATH = Path("template_companies.json")
HEADERS = ["日期", "公司名", "负责人", "项目名称", "回款金额"]


def load_template():
    if not TEMPLATE_PATH.exists(): 
        return []
    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_table():
    TABLE_PATH.parent.mkdir(parents=True, exist_ok=True)
    if TABLE_PATH.exists():
        wb = load_workbook(TABLE_PATH)
        ws = wb.active
        print(f"已加载现有表格：{TABLE_PATH}")
        print(f"工作表名称：{ws.title}")
        print(f"总行数：{ws.max_row}")
        return wb, ws
    wb, ws = Workbook(), Workbook().active
    ws.append(HEADERS)
    wb.save(TABLE_PATH)
    print(f"已创建新表格：{TABLE_PATH}")
    print(f"表头：{HEADERS}")
    return wb, ws


if __name__ == "__main__":
    wb, ws = load_table()
    print(f"操作完成！")
