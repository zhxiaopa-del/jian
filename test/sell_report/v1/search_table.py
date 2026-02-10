"""
根据明细表汇总当月数据，生成Excel汇总表。
- 第一个sheet：根据公司名进行汇总
- 第二个sheet：根据公司名和项目类型进行汇总
- 第三个sheet开始：各个公司的明细表（筛选公司，按负责人-项目类型-项目名称汇总）
运行：python search_table.py  （自动使用当前年月）
"""

import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

load_dotenv()

PROJECT_ROOT = Path(__file__).resolve().parent

# 明细表路径：data/明细表.xlsx
DETAIL_TABLE_PATH = PROJECT_ROOT / "data" / "明细表.xlsx"
TEMPLATE_PATH = PROJECT_ROOT / "template"
_DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data"


def _parse_date(val):
    """解析日期，返回(年, 月)"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.year, val.month
    try:
        dt = datetime.strptime(str(val).strip()[:10], "%Y-%m-%d")
        return dt.year, dt.month
    except ValueError:
        return None


def load_detail_rows(detail_path: Path, year: int, month: int):
    """
    从明细表读取当月数据
    明细表格式：动作、日期、公司名、负责人、项目类型、项目名称、金额、类型
    返回: [(日期, 公司名, 负责人, 项目类型, 项目名称, 金额, 类型), ...]
    """
    if not detail_path.exists():
        print(f"警告: 明细表不存在: {detail_path}")
        return []
    
    wb = load_workbook(detail_path, read_only=True, data_only=True)
    ws = wb.active
    out = []
    
    # 读取表头，确定列索引
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return []
    
    # 查找列索引
    col_map = {}
    headers = ["动作", "日期", "公司名", "负责人", "项目类型", "项目名称", "金额", "类型"]
    for idx, header in enumerate(header_row):
        if header in headers:
            col_map[header] = idx
    
    # 读取数据行
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        
        # 获取各字段值
        date_val = row[col_map.get("日期", 1)] if "日期" in col_map else None
        action = str(row[col_map.get("动作", 0)] or "").strip() if "动作" in col_map else ""
        
        # 只处理"增加"和"修改"的记录，"减少"的记录跳过
        if action == "减少":
            continue
        
        # 解析日期，只处理当前年月的数据
        pt = _parse_date(date_val)
        if not pt or pt[0] != year or pt[1] != month:
            continue
        
        # 获取其他字段
        company = str(row[col_map.get("公司名", 2)] or "").strip() if "公司名" in col_map else ""
        person = str(row[col_map.get("负责人", 3)] or "").strip() if "负责人" in col_map else ""
        project_type = str(row[col_map.get("项目类型", 4)] or "").strip() if "项目类型" in col_map else ""
        project = str(row[col_map.get("项目名称", 5)] or "").strip() if "项目名称" in col_map else ""
        type_val = str(row[col_map.get("类型", 7)] or "").strip() if "类型" in col_map else ""
        
        # 如果没有项目名称，使用项目类型
        if not project:
            project = project_type
        
        # 获取金额
        try:
            amount = float(row[col_map.get("金额", 6)] or 0) if "金额" in col_map else 0
        except (TypeError, ValueError):
            amount = 0
        
        # 只添加有金额的记录
        if amount > 0:
            out.append((date_val, company, person, project_type, project, amount, type_val))
    
    wb.close()
    return out


def aggregate_data(detail_rows):
    """
    汇总数据，区分类型（回款/合同，可能/确定/实际）：
    - by_company: 公司 -> {类型: 金额}
    - by_company_project_type: (公司, 项目类型) -> {类型: 金额}
    - detail: (公司, 负责人, 项目类型, 项目名称) -> {类型: 金额}
    """
    by_company = defaultdict(lambda: defaultdict(float))
    by_company_project_type = defaultdict(lambda: defaultdict(float))
    detail = defaultdict(lambda: defaultdict(float))
    
    for _, company, person, project_type, project, amount, type_val in detail_rows:
        # 类型字段：可能回款、确定回款、实际回款、可能合同、确定合同、实际合同等
        type_key = type_val if type_val else "实际回款"  # 默认值
        
        # 按公司+类型汇总
        if company:
            by_company[company][type_key] += amount
        
        # 按公司+项目类型+类型汇总
        if company and project_type:
            by_company_project_type[(company, project_type)][type_key] += amount
        
        # 明细：公司+负责人+项目类型+项目名称+类型
        if company and person and project_type and project:
            detail[(company, person, project_type, project)][type_key] += amount
    
    return {
        "by_company": dict(by_company),
        "by_company_project_type": dict(by_company_project_type),
        "detail": dict(detail)
    }


def _safe_set_cell(ws, row, col, value):
    """安全地设置单元格值，处理合并单元格"""
    try:
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if row >= merged_range.min_row and row <= merged_range.max_row and \
                   col >= merged_range.min_col and col <= merged_range.max_col:
                    cell = ws.cell(merged_range.min_row, merged_range.min_col)
                    break
        cell.value = value
        return True
    except Exception as e:
        try:
            ws.cell(row, col).value = value
            return True
        except:
            return False


def _safe_get_cell(ws, row, col):
    """安全地获取单元格值，处理合并单元格"""
    try:
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if row >= merged_range.min_row and row <= merged_range.max_row and \
                   col >= merged_range.min_col and col <= merged_range.max_col:
                    cell = ws.cell(merged_range.min_row, merged_range.min_col)
                    break
        return cell.value
    except:
        return None


def _find_column_index(ws, keywords, start_col=1, max_col=20):
    """查找包含关键词的列索引"""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return None
    
    for idx, header in enumerate(header_row, start=start_col):
        if idx > max_col:
            break
        header_str = str(header or "").strip()
        for keyword in keywords:
            if keyword in header_str:
                return idx
    return None


def _find_type_column(ws, type_name, start_col=1, max_col=30):
    """查找指定类型对应的列索引"""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return None
    
    for idx, header in enumerate(header_row, start=start_col):
        if idx > max_col:
            break
        header_str = str(header or "").strip()
        if header_str == type_name:
            return idx
    
    return None


def fill_sheet_1_by_company(ws, agg, year, month):
    """填充第一个sheet：根据公司名进行汇总，区分类型
    清空所有数据，取消合并单元格，只保留表结构，重新填充
    """
    # 清空所有数据并取消合并单元格
    _clear_all_data(ws)
    
    by_company = agg["by_company"]
    
    # 查找列
    col_month = _find_column_index(ws, ["月份", "月"], 1, 10) or 1
    col_company = _find_column_index(ws, ["公司"], 1, 10) or 2
    
    # 查找所有类型列
    type_columns = {}
    type_names = [
        "可能回款", "确定回款", "月初可能回款", "月初确定回款", "实际回款",
        "可能合同", "确定合同", "月初可能合同", "月初确定合同", "实际合同"
    ]
    
    for type_name in type_names:
        col_idx = _find_type_column(ws, type_name, 1, 30)
        if col_idx:
            type_columns[type_name] = col_idx
    
    # 填充数据：为每个公司创建一行
    data_start_row = 2
    for company, company_types in by_company.items():
        # 设置月份
        _safe_set_cell(ws, data_start_row, col_month, f"{month}月")
        # 设置公司名
        _safe_set_cell(ws, data_start_row, col_company, company)
        # 根据类型填充到对应列
        for type_name, amount in company_types.items():
            if type_name in type_columns:
                _safe_set_cell(ws, data_start_row, type_columns[type_name], amount)
        data_start_row += 1


def fill_sheet_2_by_company_project_type(ws, agg, year, month):
    """填充第二个sheet：根据公司名和项目类型进行汇总，区分类型
    清空所有数据，取消合并单元格，只保留表结构，重新填充
    """
    # 清空所有数据并取消合并单元格
    _clear_all_data(ws)
    
    by_company_project_type = agg["by_company_project_type"]
    
    # 查找列
    col_month = _find_column_index(ws, ["月份", "月"], 1, 10) or 1
    col_company = _find_column_index(ws, ["公司"], 1, 10) or 2
    col_project_type = _find_column_index(ws, ["项目类型", "项目分类"], 1, 10) or 3
    
    # 查找所有类型列
    type_columns = {}
    type_names = [
        "可能回款", "确定回款", "月初可能回款", "月初确定回款", "实际回款",
        "可能合同", "确定合同", "月初可能合同", "月初确定合同", "实际合同"
    ]
    
    for type_name in type_names:
        col_idx = _find_type_column(ws, type_name, 1, 30)
        if col_idx:
            type_columns[type_name] = col_idx
    
    # 填充数据：为每个(公司, 项目类型)组合创建一行
    data_start_row = 2
    for (company, project_type), project_types in by_company_project_type.items():
        # 设置月份
        _safe_set_cell(ws, data_start_row, col_month, f"{month}月")
        # 设置公司名
        _safe_set_cell(ws, data_start_row, col_company, company)
        # 设置项目类型
        _safe_set_cell(ws, data_start_row, col_project_type, project_type)
        # 根据类型填充到对应列
        for type_name, amount in project_types.items():
            if type_name in type_columns:
                _safe_set_cell(ws, data_start_row, type_columns[type_name], amount)
        data_start_row += 1


def _unmerge_all_cells(ws):
    """取消所有合并单元格"""
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))


def _clear_all_data(ws, data_start_row=2):
    """清空所有数据行（保留表头），并取消合并单元格"""
    # 先取消所有合并单元格
    _unmerge_all_cells(ws)
    
    # 从后往前删除数据行，避免索引变化
    for row_idx in range(ws.max_row, data_start_row - 1, -1):
        # 删除整行数据
        ws.delete_rows(row_idx)


def fill_sheet_company_detail(ws, detail_rows, company_name):
    """
    填充公司明细表：筛选公司，按负责人-项目类型-项目名称汇总
    只保留表结构，清空原有数据，取消合并单元格，使用明细表数据重新填充
    """
    # 先清空所有数据并取消合并单元格
    _clear_all_data(ws)
    
    # 筛选该公司的数据（公司名匹配工作表名称）
    filtered_rows = []
    for row in detail_rows:
        date_val, company, person, project_type, project, amount, type_val = row
        if company and (company_name in company or company in company_name or company == company_name):
            filtered_rows.append(row)
    
    if not filtered_rows:
        # 如果没有数据，直接返回（已经清空数据）
        return
    
    # 查找列索引（格式按照模板）
    col_date = _find_column_index(ws, ["日期"], 1, 10)
    col_company = _find_column_index(ws, ["公司", "公司名"], 1, 10) or 1
    col_person = _find_column_index(ws, ["负责人"], 1, 10) or 2
    col_project_type = _find_column_index(ws, ["项目类型", "项目分类"], 1, 10) or 3
    col_project = _find_column_index(ws, ["项目名称"], 1, 10) or 4
    
    # 查找所有类型列
    type_columns = {}
    type_names = [
        "可能回款", "确定回款", "月初可能回款", "月初确定回款", "实际回款",
        "可能合同", "确定合同", "月初可能合同", "月初确定合同", "实际合同"
    ]
    
    for type_name in type_names:
        col_idx = _find_type_column(ws, type_name, 1, 30)
        if col_idx:
            type_columns[type_name] = col_idx
    
    # 数据起始行
    data_start_row = 2
    
    # 按负责人+项目类型+项目名称分组，同时按类型汇总
    rows_by_key = defaultdict(lambda: defaultdict(float))  # (负责人, 项目类型, 项目名称) -> {类型: 金额}
    row_info = {}  # (负责人, 项目类型, 项目名称) -> (公司, 日期, 原始数据)
    
    for row_data in filtered_rows:
        date_val, company, person, project_type, project, amount, type_val = row_data
        if person and project_type and project:
            key = (person.lower(), project_type.lower(), project.lower())
            type_key = type_val if type_val else "实际回款"
            rows_by_key[key][type_key] += amount
            # 保存原始数据（使用最新的）
            row_info[key] = (company, date_val, person, project_type, project)
    
    # 按负责人分组，准备插入
    rows_by_person = defaultdict(list)
    for key, type_amounts in rows_by_key.items():
        company, date_val, person, project_type, project = row_info[key]
        rows_by_person[person].append((company, date_val, project_type, project, type_amounts))
    
    # 插入新数据
    insert_row = data_start_row
    
    for person, items in rows_by_person.items():
        # 处理该负责人的所有项目
        for company, date_val, project_type, project, type_amounts in items:
            # 添加新行：使用明细表的数据
            if col_company and col_company > 0:
                _safe_set_cell(ws, insert_row, col_company, company)
            _safe_set_cell(ws, insert_row, col_person, person)
            _safe_set_cell(ws, insert_row, col_project_type, project_type)
            _safe_set_cell(ws, insert_row, col_project, project)
            if col_date and col_date > 0:
                _safe_set_cell(ws, insert_row, col_date, date_val)
            
            # 根据类型填充到对应列
            for type_key, amount in type_amounts.items():
                if type_key in type_columns:
                    _safe_set_cell(ws, insert_row, type_columns[type_key], amount)
            
            insert_row += 1


def _remove_empty_and_duplicate_rows(ws, col_person, col_project_type, col_project, data_start_row=2):
    """删除空行和重复行"""
    rows_to_delete = []
    seen_keys = set()
    
    for row_idx in range(data_start_row, ws.max_row + 1):
        person = str(_safe_get_cell(ws, row_idx, col_person) or "").strip()
        project_type = str(_safe_get_cell(ws, row_idx, col_project_type) or "").strip()
        project = str(_safe_get_cell(ws, row_idx, col_project) or "").strip()
        
        # 跳过汇总行
        if person in ["小计", "合计", "汇总"]:
            continue
        
        # 检查空行
        is_empty = not (person or project_type or project)
        
        # 检查重复行
        is_duplicate = False
        if person and project_type and project:
            key = (person.lower(), project_type.lower(), project.lower())
            if key in seen_keys:
                is_duplicate = True
            else:
                seen_keys.add(key)
        
        if is_empty or is_duplicate:
            rows_to_delete.append(row_idx)
    
    # 从后往前删除
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)


def get_template_file():
    """获取模板文件"""
    template_dir = TEMPLATE_PATH
    if not template_dir.exists():
        print(f"错误: 模板目录不存在: {template_dir}")
        return None
    
    xlsx_files = sorted([p for p in template_dir.glob("*.xlsx") if not p.name.startswith("~$")])
    if not xlsx_files:
        print(f"错误: 模板目录中没有找到Excel文件: {template_dir}")
        return None
    
    return xlsx_files[0]


def fill_template(template_path: Path, output_path: Path, detail_rows, agg: dict, year: int, month):
    """填充模板并生成汇总表"""
    # 处理已存在的文件
    if output_path.exists():
        try:
            output_path.unlink()
        except PermissionError:
            backup_path = output_path.with_suffix('.bak.xlsx')
            try:
                if backup_path.exists():
                    backup_path.unlink()
                output_path.rename(backup_path)
                print(f"已备份原文件到: {backup_path}")
            except Exception as e:
                print(f"错误: 无法处理已存在的文件: {str(e)}")
                return
    
    # 复制模板
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    
    sheet_names = wb.sheetnames
    
    # 第一个sheet：根据公司名进行汇总
    if len(sheet_names) > 0:
        fill_sheet_1_by_company(wb[sheet_names[0]], agg, year, month)
    
    # 第二个sheet：根据公司名和项目类型进行汇总
    if len(sheet_names) > 1:
        fill_sheet_2_by_company_project_type(wb[sheet_names[1]], agg, year, month)
    
    # 第三个sheet开始：各个公司的明细表
    for i in range(2, len(sheet_names)):
        ws = wb[sheet_names[i]]
        sheet_name = sheet_names[i]
        fill_sheet_company_detail(ws, detail_rows, sheet_name)
        
        # 删除空行和重复行
        col_person = _find_column_index(ws, ["负责人"], 1, 10) or 2
        col_project_type = _find_column_index(ws, ["项目类型", "项目分类"], 1, 10) or 3
        col_project = _find_column_index(ws, ["项目名称"], 1, 10) or 4
        
        if col_person and col_project_type and col_project:
            _remove_empty_and_duplicate_rows(ws, col_person, col_project_type, col_project)
    
    wb.save(output_path)
    wb.close()


def main():
    now = datetime.now()
    year, month = now.year, now.month

    # 从明细表读取当月数据
    detail_rows = load_detail_rows(DETAIL_TABLE_PATH, year, month)
    
    if not detail_rows:
        print(f"警告: {year}年{month}月没有找到数据")
        return

    # 汇总数据
    agg = aggregate_data(detail_rows)
    
    # 获取模板文件
    template_file = get_template_file()
    if not template_file:
        return
    
    # 生成输出文件
    out_path = _DEFAULT_OUTPUT_DIR / str(year) / f"{month}月份回款汇总.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 填充模板并生成汇总表
    fill_template(template_file, out_path, detail_rows, agg, year, month)
    
    # 计算总金额（所有公司所有类型的总和）
    total = 0
    for company_types in agg["by_company"].values():
        if isinstance(company_types, dict):
            total += sum(company_types.values())
        else:
            total += company_types
    
    print(f"已汇总 {year}年{month}月 回款 {len(detail_rows)} 条，合计 {total}，已写入 {out_path}")


if __name__ == "__main__":
    main()
