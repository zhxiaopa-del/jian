"""
自动识别回款汇总配置。
"""
import json
from pathlib import Path
from typing import Dict, Optional, Any
from openpyxl import load_workbook

SUMMARY_CONFIG_FILENAME = "summary_config.json"


def _find_col_actual(ws, max_row: int = 3, max_col: int = 15) -> Optional[int]:
    """在表头前几行中找「实际回款」列索引（0-based），未找到返回 None"""
    for r in range(1, min(max_row + 1, ws.max_row + 1)):
        for c in range(1, min(max_col, ws.max_column + 1)):
            cell_value = ws.cell(row=r, column=c).value
            if cell_value and "实际回款" in str(cell_value).strip():
                return c - 1
    return None


def _save_json(path: Path, data: Any):
    """保存JSON文件"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _detect_summary_config(project_root: Path) -> Dict[str, Any]:
    """从 data/、template/ 自动识别路径、表头、实际回款列、分公司 sheet、项目分类"""
    config = {}
    data_dir = project_root / "data"
    template_dir = project_root / "template"
    
    # 回款表路径
    xlsx_list = sorted([p for p in data_dir.glob("*.xlsx") if not p.name.startswith("~$")])
    non_summary = [p for p in xlsx_list if "汇总" not in p.stem]
    if non_summary:
        config["table_path"] = str(non_summary[0].resolve().relative_to(project_root))
    elif xlsx_list:
        config["table_path"] = str(xlsx_list[0].resolve().relative_to(project_root))
    
    # 表头
    if config.get("table_path"):
        table_full = project_root / config["table_path"]
        try:
            wb = load_workbook(table_full, read_only=True, data_only=True)
            row1 = next(wb.active.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()
            if row1 and any(row1):
                config["headers"] = [str(c).strip() if c else "" for c in row1]
        except (PermissionError, Exception) as e:
            print(f"警告: 无法读取表头 {table_full}: {str(e)}")
    
    # 模板路径
    template_xlsx = sorted([p for p in template_dir.glob("*.xlsx") if not p.name.startswith("~$")])
    if template_xlsx:
        config["template_path"] = str(template_xlsx[0].resolve().relative_to(project_root))
    
    # 从模板识别配置
    if config.get("template_path"):
        template_full = project_root / config["template_path"]
        try:
            wb = load_workbook(template_full, read_only=True, data_only=True)
            all_sheets = wb.sheetnames
            summary_sheets = [s for s in all_sheets if "汇总" in s]
            config["branch_sheet_names"] = [s for s in all_sheets if s not in summary_sheets]
            
            # 实际回款列
            summary_total = next((s for s in summary_sheets if "总" in s and "详细" not in s), summary_sheets[0] if summary_sheets else None)
            summary_detail = next((s for s in summary_sheets if "详细" in s), summary_sheets[1] if len(summary_sheets) > 1 else summary_sheets[0] if summary_sheets else None)
            
            if summary_total:
                col = _find_col_actual(wb[summary_total])
                if col is not None:
                    config["col_actual_total"] = col
            
            if summary_detail:
                col = _find_col_actual(wb[summary_detail])
                if col is not None:
                    config["col_actual_detail"] = col
            
            if config.get("branch_sheet_names"):
                col = _find_col_actual(wb[config["branch_sheet_names"][0]], max_row=2)
                if col is not None:
                    config["col_actual_branch"] = col
            
            # 项目分类
            skip = {"小计", "合计", "汇总", "项目分类", "负责人", "项目名称", "项目类型"}
            projects = []
            seen = set()
            for sn in summary_sheets:
                ws = wb[sn]
                for r in range(2, min(30, ws.max_row + 1)):
                    s = (ws.cell(row=r, column=2).value or "").strip()
                    if s and s not in skip and s not in seen:
                        seen.add(s)
                        projects.append(s)
                if projects:
                    break
            
            if projects:
                config["project_order"] = projects
            
            wb.close()
        except (PermissionError, Exception) as e:
            print(f"警告: 无法读取模板文件 {template_full}: {str(e)}")
    
    return config


def save_summary_config(config_path: Optional[str] = None, project_root: Optional[str] = None) -> Dict[str, Any]:
    """检测并保存 summary_config.json，返回 config"""
    root = (Path(project_root) if project_root else Path(__file__).resolve().parent).resolve()
    path = Path(config_path or root / SUMMARY_CONFIG_FILENAME)
    config = _detect_summary_config(root)
    _save_json(path, config)
    print(f"已自动识别并保存配置到 {path}")
    return config


def load_summary_config(config_path: Optional[str] = None, project_root: Optional[str] = None) -> Dict[str, Any]:
    """加载 summary_config.json；不存在则返回本次检测结果"""
    root = (Path(project_root) if project_root else Path(__file__).resolve().parent).resolve()
    path = Path(config_path or root / SUMMARY_CONFIG_FILENAME)
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return _detect_summary_config(root)
