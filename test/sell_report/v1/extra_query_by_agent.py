import json
import re
from datetime import datetime
from pathlib import Path
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ================= 配置信息 =================
BASE_URL = "http://10.3.0.16:8100/v1" 
API_KEY = "222442bb160d5081b9e38506901d6889"  
MODEL = "qwen3-14b"     

client = OpenAI(
    api_key=API_KEY,
    base_url=BASE_URL,
    timeout=60.0,
)

class DataExtractor:
    def __init__(self):
        self.valid_types = [
            "可能回款", "确定回款", "月初可能回款", "月初确定回款", "实际回款",
            "可能合同", "确定合同", "月初可能合同", "月初确定合同", "实际合同"
        ]
        

    def extract_fields(self, user_input):
        # 构造更全面的 Prompt
        system_prompt = f"""你是一个专业的数据提取助手。请从用户输入提取信息，输出 JSON。

【规则】：
1. 动作（必填字段）: 
   - 这是最重要的字段，必须从以下三个选项中选择一个："增加"、"减少"、"修改"。
   - "增加"：用户描述的是新增记录、新发生的事件
     * 关键词：收到了、签了、新增、添加、完成了、到账了、确认了、达成了等
     * 示例："收到了5万元回款"、"签了合同"、"新增一个项目"
   - "减少"：用户描述的是删除、取消、撤销记录
     * 关键词：删除了、取消了、撤销了、作废了、退回了、取消了等
     * 示例："删除了这条记录"、"取消了合同"、"撤销了回款"
   - "修改"：用户描述的是更新、修改、调整现有记录
     * 关键词：修改了、更新了、改成了、调整了、变更了、更正了等
     * 示例："修改了金额"、"更新了日期"、"改成了确定回款"
   - 如果无法明确判断，默认选择"增加"。
2. 日期: 
   - 必须是 "YYYY-MM-DD" 格式。
   - 你需要根据当前日期理解用户提到的相对日期（如"今天"、"昨天"、"2月1号"等）。
   - 如果用户提到"昨天"，你需要计算出昨天的日期。
   - 如果用户只说了月份和日期（如"2月1号"、"2月1日"），你需要判断年份（通常是今年，但如果已经过了这个日期且接近年底，可能是明年）。
   - 务必将各种日期表示（如"2月1日"、"2月1号"、"今天"、"昨天"）转换为标准的 "YYYY-MM-DD" 格式。
   - 如果用户没有提到日期，默认使用今天的日期。
3. 项目分类（必填字段）:
   - 从用户输入中识别项目的分类或类别。
   - 常见的项目分类包括：环保、维护、软件、咨询、IT运维、动环项目、地铁测温、煤矿项目、污水处理、咨询服务等。
   - 如果用户明确提到了项目分类（如"环保项目"、"IT运维项目"、"软件项目"），提取该类型。
   - 如果用户提到了具体的项目名称，可以根据项目名称推断项目分类。
   - 如果无法识别，可以留空或根据上下文推断。
4. 项目名称（必填字段）:
   - 从用户输入中提取具体的项目名称。
   - 项目名称通常是具体的业务项目名称，如"污水处理项目"、"IT运维项目"、"咨询服务项目"等。
   - 如果用户提到了项目名称，直接提取。
   - 如果用户只提到了项目分类但没有具体名称，可以结合公司名和项目分类生成一个合理的项目名称。
   - 如果无法识别，可以留空。
5. 金额: 提取为数字。
6. 类型: 从 {self.valid_types} 中选择。

【输出示例】：
{{"动作": "增加", "日期": "2024-02-01", "公司名": "...", "负责人": "...", "项目分类": "环保", "项目名称": "污水处理项目", "金额": 0, "类型": "实际回款"}}
"""

        try:
            response = client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"用户输入：{user_input}"}
                ],
                temperature=0,
            )

            content = response.choices[0].message.content.strip()
            json_match = re.search(r"\{.*\}", content, re.DOTALL)
            if json_match:
                data = json.loads(json_match.group())
                return data
            else:
                return {"error": "JSON解析失败", "raw": content}

        except Exception as e:
            return {"error": f"调用失败: {str(e)}"}
    
    def save_to_excel(self, data_list, output_path=None):
        """
        将提取的数据列表保存到Excel文件，如果文件存在则更新（追加数据）
        
        Args:
            data_list: 数据列表，每个元素是一个字典
            output_path: 输出文件路径，如果为None则在data文件夹下创建明细表.xlsx
        """
        if not data_list:
            print("警告: 数据列表为空，无法保存")
            return None
        
        # 确定输出路径：在data文件夹下创建明细表.xlsx
        if output_path is None:
            # 获取当前文件所在目录
            current_dir = Path(__file__).parent
            data_dir = current_dir / "data"
            data_dir.mkdir(exist_ok=True)  # 如果data文件夹不存在则创建
            output_path = data_dir / "明细表.xlsx"
        else:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 定义表头
        headers = ["动作", "日期", "公司名", "负责人", "项目分类", "项目名称", "金额", "类型"]
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 如果文件存在，读取现有数据
        existing_data = []
        if output_path.exists():
            try:
                wb = load_workbook(output_path)
                ws = wb.active
                
                # 读取现有数据（跳过表头）
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):  # 跳过空行
                        row_dict = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
                        existing_data.append(row_dict)
                
                wb.close()
                print(f"读取到 {len(existing_data)} 条现有数据")
            except Exception as e:
                print(f"警告: 读取现有文件失败，将创建新文件: {str(e)}")
                existing_data = []
        
        # 合并现有数据和新数据，去重
        # 使用日期+公司名+负责人+项目名称+类型作为唯一标识
        data_dict = {}
        
        # 先添加现有数据
        for item in existing_data:
            key = (
                str(item.get("日期", "")),
                str(item.get("公司名", "")),
                str(item.get("负责人", "")),
                str(item.get("项目名称", "")),
                str(item.get("类型", ""))
            )
            data_dict[key] = item
        
        # 添加新数据（跳过错误数据）
        new_count = 0
        update_count = 0
        for data in data_list:
            if "error" in data:
                continue
            
            key = (
                str(data.get("日期", "")),
                str(data.get("公司名", "")),
                str(data.get("负责人", "")),
                str(data.get("项目名称", "")),
                str(data.get("类型", ""))
            )
            
            if key in data_dict:
                # 如果已存在，根据动作决定是更新还是删除
                action = data.get("动作", "增加")
                if action == "减少":
                    # 删除记录
                    del data_dict[key]
                    print(f"删除记录: {key}")
                elif action == "修改":
                    # 更新记录
                    data_dict[key] = data
                    update_count += 1
                    print(f"更新记录: {key}")
                # 如果是"增加"且已存在，不重复添加
            else:
                # 新增记录
                if data.get("动作", "增加") != "减少":  # 减少操作且不存在记录时，不添加
                    data_dict[key] = data
                    new_count += 1
        
        # 创建或更新Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "明细表"
        
        # 写入表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, data in enumerate(data_dict.values(), start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = data.get(header, "")
                # 处理空值
                if value is None:
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 设置列宽
        column_widths = {
            "A": 12,  # 动作
            "B": 15,  # 日期
            "C": 20,  # 公司名
            "D": 15,  # 负责人
            "E": 15,  # 项目分类
            "F": 25,  # 项目名称
            "G": 15,  # 金额
            "H": 15,  # 类型
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置数据行对齐方式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 保存文件
        wb.save(output_path)
        total_count = len(data_dict)
        print(f"数据已保存到: {output_path}")
        print(f"统计: 新增 {new_count} 条，更新 {update_count} 条，总计 {total_count} 条记录")
        return str(output_path)

# ================= 测试 =================
if __name__ == "__main__":
    extractor = DataExtractor()
    
    test_cases = [
        "丁辉说2月1号上海环保公司那个污水处理项目收到了5.5万元实际回款",
        "昨天北京联想的IT运维项目确定回款10万",
        "三源公司的咨询服务项目已经签了，合同金额20000"
    ]

    results = []
    for text in test_cases:
        print(f"输入: {text}")
        result = extractor.extract_fields(text)
        print("提取结果:")
        print(json.dumps(result, indent=4, ensure_ascii=False))
        results.append(result)
        print("-" * 50)
    
    # # 保存到Excel
    # if results:
    #     extractor.save_to_excel(results)
