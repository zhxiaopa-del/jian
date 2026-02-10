"""
按当前月份回款汇总：从回款表统计当月数据，按公司/负责人/项目汇总后填入 template 样式 Excel。
运行：python feb_summary.py  （自动使用当前年月）
"""

import json
import os
import re
import shutil
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook


load_dotenv()

PROJECT_ROOT = Path(__file__).resolve().parent
_config = load_summary_config(project_root=PROJECT_ROOT)


def _path(p):
    if not p:
        return None
    path = Path(p)
    return path if path.is_absolute() else (PROJECT_ROOT / path)


TABLE_PATH = _path(_config.get("table_path"))
TEMPLATE_PATH = _path(_config.get("template_path"))
_DEFAULT_OUTPUT_DIR = (TABLE_PATH.parent if TABLE_PATH and TABLE_PATH.parent else PROJECT_ROOT / "data")

COL_ACTUAL_TOTAL = _config.get("col_actual_total", 6)
COL_ACTUAL_DETAIL = _config.get("col_actual_detail", 7)
COL_ACTUAL_BRANCH = _config.get("col_actual_branch", 7)
BRANCH_SHEET_NAMES = _config.get("branch_sheet_names") or []
PROJECT_ORDER = _config.get("project_order") or []


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.year, val.month
    try:
        dt = datetime.strptime(str(val).strip()[:10], "%Y-%m-%d")
        return dt.year, dt.month
    except ValueError:
        return None


def load_payment_rows(table_path: Path, year: int, month: int):
    wb = load_workbook(table_path, read_only=True, data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        pt = _parse_date(row[0])
        if not pt or pt[0] != year or pt[1] != month:
            continue
        try:
            amount = float(row[4] or 0)
        except (TypeError, ValueError):
            amount = 0
        out.append((row[0], (row[1] or "").strip(), (row[2] or "").strip(), (row[3] or "").strip(), amount))
    wb.close()
    return out


def normalize_project(name):
    if not name:
        return ""
    n = name.strip()
    for p in PROJECT_ORDER:
        if p in n or n in p:
            return p
    return n


def aggregate(rows):
    detail, by_person, by_company, by_cp = defaultdict(float), defaultdict(float), defaultdict(float), defaultdict(float)
    for _, company, person, project, amount in rows:
        proj = normalize_project(project) or project or "其他"
        detail[(company, person, proj)] += amount
        by_person[(company, person)] += amount
        by_company[company] += amount
        csum = (company + "汇总") if company and not company.endswith("汇总") else company
        if csum:
            by_cp[(csum, proj)] += amount
        by_cp[("小计", proj)] += amount
    by_cp[("合计", None)] = sum(by_company.values())
    return {"detail": dict(detail), "by_person": dict(by_person), "by_company": dict(by_company), "by_company_project": dict(by_cp)}


def align_with_llm(rows_data, template_ref):
    try:
        from langchain_community.chat_models import ChatTongyi
    except ImportError:
        return rows_data
    if not os.getenv("DASHSCOPE_API_KEY"):
        return rows_data
    lines = "\n".join(f"  {c}\t{p}\t{proj}\t{amt}" for c, p, proj, amt in rows_data)
    prompt = f"回款明细（公司\\t负责人\\t项目\\t金额）：\n{lines}\n模板参考：{json.dumps(template_ref[:30], ensure_ascii=False)}\n请将公司名、负责人、项目名校准为与模板一致，只输出JSON数组 [{{\"公司名\":\"\",\"负责人\":\"\",\"项目名称\":\"\",\"金额\":0}}]"
    res = ChatTongyi(model_name="qwen-turbo").invoke(prompt).content
    m = re.search(r"\[[\s\S]*?\]", res)
    if not m:
        return rows_data
    try:
        arr = json.loads(m.group())
        out = [(str(i.get("公司名") or "").strip(), str(i.get("负责人") or "").strip(), str(i.get("项目名称") or "").strip(), float(i.get("金额") or 0)) for i in arr if i]
        return out if out else rows_data
    except (json.JSONDecodeError, TypeError, ValueError):
        return rows_data


def _fill_sheet_total(ws, agg):
    by_cp, by_company = agg["by_company_project"], agg["by_company"]
    block = None
    for row in range(4, ws.max_row + 1):
        label = (ws.cell(row, 1).value or "").strip()
        proj = (ws.cell(row, 2).value or "").strip() or None
        if label and label.endswith("汇总"):
            block = label
        if (label, proj) in by_cp:
            ws.cell(row, COL_ACTUAL_TOTAL + 1, value=by_cp[(label, proj)])
        elif label == "小计" and not proj and block:
            ws.cell(row, COL_ACTUAL_TOTAL + 1, value=by_company.get(block.replace("汇总", ""), 0))
        elif label == "合计" and not proj:
            ws.cell(row, COL_ACTUAL_TOTAL + 1, value=by_cp.get(("合计", None), 0))


def _fill_sheet_detail(ws, agg):
    detail, by_cp, by_company = agg["detail"], agg["by_company_project"], agg["by_company"]
    cur_company, cur_person = "", ""
    for row in range(2, ws.max_row + 1):
        company = (ws.cell(row, 1).value or "").strip() or cur_company
        person = (ws.cell(row, 2).value or "").strip() or cur_person
        proj = (ws.cell(row, 3).value or "").strip()
        if company:
            cur_company = company
        if person:
            cur_person = person
        val = detail.get((company, person, proj)) if proj else None
        if val is None and person and person.endswith("汇总") and proj:
            val = by_cp.get((person, proj), 0)
        if val is None and person == "小计" and not proj and cur_company:
            val = by_company.get(cur_company.replace("汇总", ""), 0)
        if val is not None:
            ws.cell(row, COL_ACTUAL_DETAIL + 1, value=val)


def _fill_sheet_branch(ws, company, agg):
    detail, by_cp = agg["detail"], agg["by_company_project"]
    cur_person, cur_proj = "", ""
    for row in range(3, ws.max_row + 1):
        s1 = (ws.cell(row, 1).value or "").strip()
        s2 = (ws.cell(row, 2).value or "").strip()
        s3 = (ws.cell(row, 3).value or "").strip()
        if s1 and "预计合同" in s1:
            break
        if s1 == "负责人" and row > 3:
            break
        if s1:
            cur_person = s1
        if s2 and s2 in PROJECT_ORDER:
            cur_proj = s2
        val = None
        if s3 == "小计":
            val = detail.get((company, cur_person, cur_proj), 0)
        elif cur_person and cur_person.endswith("汇总") and s2 in PROJECT_ORDER:
            val = by_cp.get((cur_person, s2), 0)
        elif (cur_person or s1) and s2 in PROJECT_ORDER:
            val = detail.get((company, cur_person or s1, s2), 0)
        if val is not None:
            ws.cell(row, COL_ACTUAL_BRANCH + 1, value=val)


def fill_template(template_path: Path, output_path: Path, agg: dict):
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    _fill_sheet_total(wb["预计汇总（总）"], agg)
    _fill_sheet_detail(wb["预计汇总（详细）"], agg)
    for sn in BRANCH_SHEET_NAMES:
        _fill_sheet_branch(wb[sn], sn, agg)
    wb.save(output_path)
    wb.close()


def load_template_companies():
    with open(PROJECT_ROOT / "template_companies.json", encoding="utf-8") as f:
        return json.load(f)


def main():
    now = datetime.now()
    year, month = now.year, now.month

    rows = load_payment_rows(TABLE_PATH, year, month)  # 无记录时为空列表，汇总结果为 0

    if os.getenv("DASHSCOPE_API_KEY"):
        merged = defaultdict(float)
        for _, c, p, proj, amt in rows:
            merged[(c, p, normalize_project(proj))] += amt
        aligned = align_with_llm([(*k, v) for k, v in merged.items()], load_template_companies())
        if aligned:
            rows = [(date(year, month, 1), c, p, proj, amt) for c, p, proj, amt in aligned]

    agg = aggregate(rows)
    out_path = _DEFAULT_OUTPUT_DIR / str(year) / f"{month}月份回款汇总.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fill_template(TEMPLATE_PATH, out_path, agg)
    total = agg["by_company_project"].get(("合计", None), 0)
    print(f"已汇总 {year}年{month}月 回款 {len(rows)} 条，合计 {total}，已写入 {out_path}")


if __name__ == "__main__":
    main()
