import json, os, re
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from langchain_community.chat_models import ChatTongyi
from langchain_core.messages import HumanMessage, AIMessage

load_dotenv()

TABLE_PATH = Path("data/回款表.xlsx")
TEMPLATE_PATH = Path("template_companies.json")
HEADERS = ["日期", "公司名", "负责人", "项目分类", "项目名称", "回款金额"]

class PaymentAgent:
    def __init__(self):
        self.llm = ChatTongyi(model_name="qwen-turbo")
        self.wb, self.ws = self._load_table()
        self.template_data = self._load_template()
        self.state = "IDLE" 
        self.pending = None     # 暂存正在提取的信息
        self.choices = []       # 暂存待选列表
        self.history = []       # 对话上下文

    def _load_template(self):
        if not TEMPLATE_PATH.exists(): return []
        with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)

    def _load_table(self):
        TABLE_PATH.parent.mkdir(exist_ok=True)
        if TABLE_PATH.exists():
            wb = load_workbook(TABLE_PATH)
            return wb, wb.active
        wb, ws = Workbook(), Workbook().active
        ws.append(HEADERS)
        wb.save(TABLE_PATH)
        print(wb,ws)
        return wb, ws

    def _ask_llm(self, prompt, use_history=True):
        messages = []
        if use_history:
            messages.extend(self.history[-5:]) # 保留最近5轮
        messages.append(HumanMessage(content=prompt))
        
        res = self.llm.invoke(messages).content
        # 更新记忆
        self.history.append(HumanMessage(content=prompt))
        self.history.append(AIMessage(content=res))
        
        match = re.search(r"\{.*\}", res, re.DOTALL)
        return json.loads(match.group()) if match else {}

    def _llm_confirm_intent(self, user_text, context):
        """用大模型判断用户是「确认」还是「取消」，返回 True=确认，False=取消"""
        prompt = f"""语境：{context}
用户回复：「{user_text}」
请判断用户是在确认（同意/要执行）还是取消（不同意/不做了）。只输出JSON：{{"intent":"确认"}} 或 {{"intent":"取消"}}，不要其他内容。"""
        out = self._ask_llm(prompt, use_history=False)
        return (out.get("intent") or "").strip() == "确认"

    def _fuzzy_match_person(self, user_person, template_person):
        """负责人模糊匹配：完全一致、包含、或编辑距离为1（如丁辉→丁俊）"""
        if not user_person or not template_person:
            return False
        if user_person == template_person:
            return True
        if user_person in template_person or template_person in user_person:
            return True
        # 编辑距离为 1 时也认为匹配（如丁辉 vs 丁俊）
        n, m = len(user_person), len(template_person)
        if abs(n - m) > 1:
            return False
        # Levenshtein distance == 1
        if n > m:
            n, m, user_person, template_person = m, n, template_person, user_person
        for i in range(n):
            if user_person[i] != template_person[i]:
                return user_person[i + 1:] == template_person[i + 1:] or user_person[i:] == template_person[i + 1:]
        return n + 1 == m

    def _get_template_candidates(self, extracted):
        """根据已有信息从模板筛选；支持负责人模糊匹配；结果按(公司名,项目名称)去重"""
        c = (extracted.get("公司名") or "").strip()
        p = (extracted.get("负责人") or "").strip()
        proj = (extracted.get("项目名称") or "").strip()

        raw = []
        for item in self.template_data:
            match_company = (not c) or (c in item["公司名"]) or (item["公司名"] in c)
            match_person = (not p) or self._fuzzy_match_person(p, item["负责人"])
            match_proj = (not proj) or (proj in item["项目名称"]) or (item["项目名称"] in proj)
            if match_person and (match_company or match_proj):
                raw.append(item)
        if not raw and p:
            # 仅凭负责人再试一次：只做负责人模糊匹配
            for item in self.template_data:
                if self._fuzzy_match_person(p, item["负责人"]):
                    raw.append(item)
        if not raw:
            return []

        # 按 (公司名, 项目名称) 去重，保留一条代表
        seen = set()
        results = []
        for item in raw:
            key = (item["公司名"], item["项目名称"])
            if key not in seen:
                seen.add(key)
                results.append(item)
        return results

    def handle_idle(self, text):
        """解析输入并寻找匹配项"""
        prompt = f"""解析回款信息。当前日期：{datetime.now().strftime("%Y-%m-%d")}
        用户输入：{text}
        请提取：日期、公司名、负责人、项目名称、回款金额。
        输出JSON格式：{{"日期":"","公司名":"","负责人":"","项目名称":"","回款金额":0}}
        注意：如果没有提到日期，默认使用今天。金额必须是纯数字。"""
        
        extracted = self._ask_llm(prompt)
        if not extracted.get("回款金额"):
            return "未能识别到有效金额，请说明回款多少钱。"

        # 匹配模板
        candidates = self._get_template_candidates(extracted)
        
        if len(candidates) == 1:
            # 完美匹配
            self.pending = {**extracted, **candidates[0]}
            self.state = "CONFIRM_INFO"
            return self._format_confirm_msg()
        
        elif len(candidates) > 1:
            # 存在多个可能
            self.choices = candidates
            self.pending = extracted
            self.state = "WAIT_CHOICE"
            msg = "查找到多个匹配的项目，请选择编号（输入1, 2...）：\n"
            for i, cand in enumerate(candidates, 1):
                msg += f"{i}. 【{cand['公司名']}】{cand['负责人']} - {cand['项目名称']}\n"
            return msg
        
        else:
            return "未在模板中找到对应的公司、负责人或项目，请检查后重新输入。"

    def handle_wait_choice(self, text):
        """处理用户的编号选择"""
        if text.isdigit() and 1 <= int(text) <= len(self.choices):
            selected = self.choices[int(text)-1]
            self.pending.update(selected)
            self.state = "CONFIRM_INFO"
            return self._format_confirm_msg()
        else:
            # 如果不是数字，尝试重新解析（可能用户直接说了项目名）
            return self.handle_idle(text)

    def handle_confirm_info(self, text):
        """确认并检查重复（用大模型判断用户是确认还是取消）"""
        if not self._llm_confirm_intent(text, "刚才向用户展示了回款信息，询问是否确认录入。"):
            self.state = "IDLE"
            return "已取消，请重新输入。"

        # 检查近7天内：同一公司、同一负责人、同一项目、同一回款金额 才提示重复
        amount = float(self.pending["回款金额"])
        company = (self.pending.get("公司名") or "").strip()
        person = (self.pending.get("负责人") or "").strip()
        project = (self.pending.get("项目名称") or "").strip()
        duplicates = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            try:
                r_date = row[0] if isinstance(row[0], datetime) else datetime.strptime(str(row[0])[:10], "%Y-%m-%d")
                if (datetime.now() - r_date).days > 7:
                    continue
                same_amount = abs(float(row[4]) - amount) < 0.1
                same_company = (row[1] or "").strip() == company
                same_person = (row[2] or "").strip() == person
                same_project = (row[3] or "").strip() == project
                if same_amount and same_company and same_person and same_project:
                    duplicates.append(f"{row[0]} {row[1]} {row[2]} {row[3]} {row[4]}元")
            except: continue

        if duplicates:
            self.state = "CONFIRM_DUP"
            return "⚠️ 近7天有相同金额记录：\n" + "\n".join(duplicates) + "\n是否确定仍要重复添加？"
        
        return self.execute_save()

    def handle_confirm_dup(self, text):
        """用大模型判断用户是否仍要添加（如「添加」「确定」「要」等均为确认）"""
        if self._llm_confirm_intent(text, "系统提示近7天有相同金额记录，问用户是否确定仍要重复添加。"):
            return self.execute_save()
        self.state = "IDLE"
        return "已取消录入。"

    def _format_confirm_msg(self):
        p = self.pending
        return (f"请确认信息：\n📅 日期：{p['日期']}\n🏢 公司：{p['公司名']}\n👤 负责人：{p['负责人']}\n"
                f"🏗️ 项目：{p['项目名称']}\n💰 金额：{p['回款金额']}\n确认请回复‘确定’。")

    def execute_save(self):
        p = self.pending
        self.ws.append([p["日期"], p["公司名"], p["负责人"], p["项目名称"], p["回款金额"]])
        self.wb.save(TABLE_PATH)
        self.state = "IDLE"
        return f"✅ 录入成功！已存入表格。"

    def run(self):
        print(f"回款 Agent 已启动。上下文记忆已开启。")
        handlers = {
            "IDLE": self.handle_idle,
            "WAIT_CHOICE": self.handle_wait_choice,
            "CONFIRM_INFO": self.handle_confirm_info,
            "CONFIRM_DUP": self.handle_confirm_dup
        }
        
        while True:
            user_input = input("\n您：").strip()
            if user_input.lower() in ["exit", "退出"]: break
            
            handler = handlers.get(self.state, self.handle_idle)
            response = handler(user_input)
            print(f"Agent：{response}")

if __name__ == "__main__":
    PaymentAgent().run()