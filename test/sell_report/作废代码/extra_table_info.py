"""
从模板 Excel 提取公司信息，并自动识别回款汇总配置。
主函数：extra_info() - 按照5个步骤执行，可封装为 MCP。
"""
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Any
from dotenv import load_dotenv
from openpyxl import load_workbook
from langchain_community.chat_models import ChatTongyi

load_dotenv()

EXTRACT_PROMPT = """你是一个表格分析助手。下面是从 Excel 导出的多个工作表（每行用制表符分隔列）。

请按以下步骤分析并提取：

第一步：分析表格结构
- 第二个表：已标注为「预计汇总」，其中包含「项目名称」或「项目」或「项目分类」等列，以及公司名、负责人（或可对应到第一个表的列）。

第二步：对应关系
- 根据「公司名」「负责人」将两个表对应起来：第一个表中的每条「公司名+负责人」，在第二个表「预计汇总」中找到同一公司名、同一负责人的那一行，读取该行的「项目名称」/「项目」/「项目分类」列的值。

第三步：输出
- 为每条「公司名+负责人」输出一条记录，包含：公司名、负责人、项目名称（从「预计汇总」中分析得到的项目名称；若无法对应则留空）。
- 去重：公司名+负责人+项目名称完全相同的只保留一条。

最终只输出一个 JSON 数组，不要任何其他说明或 markdown 标记。格式严格如下：

[{{"公司名": "xxx", "负责人": "xxx", "项目类型": "xxx", "项目名称": "xxx"}}, ...]

表格内容：
{table_text}
"""

SUMMARY_CONFIG_FILENAME = "summary_config.json"
TEMPLATE_COMPANIES_FILENAME = "template_companies.json"


def read_xlsx(dir_path: str) -> str:
    """读取Excel并返回文本"""
    all_texts = []
    for p in sorted(Path(dir_path).glob("*.xlsx")):
        # 跳过Excel临时锁定文件（以~$开头）
        if p.name.startswith("~$"):
            continue
        
        try:
            wb = load_workbook(p, data_only=True, read_only=True)
            for i, ws in enumerate(wb.worksheets):
                rows = [[str(c) if c is not None else "" for c in r] for r in ws.iter_rows(values_only=True)]
                is_summary = (i == 1 or "预计" in ws.title)
                label = "【第二个表：预计汇总（此处含项目名称列，请分析并对应到公司名、负责人）】" if is_summary else f"【工作表: {ws.title}】"
                content = "\n".join(["\t".join(r) for r in rows])
                all_texts.append(f"{label}\n{content}")
            wb.close()
        except PermissionError:
            # 如果文件被占用，跳过该文件
            print(f"警告: 文件被占用，跳过: {p.name}")
            continue
        except Exception as e:
            # 其他错误也跳过，避免中断整个流程
            print(f"警告: 读取文件失败 {p.name}: {str(e)}")
            continue
    return "\n\n".join(all_texts)


def _find_col_actual(ws, max_row: int = 3, max_col: int = 15) -> Optional[int]:
    """在表头前几行中找「实际回款」列索引（0-based），未找到返回 None"""
    for r in range(1, min(max_row + 1, ws.max_row + 1)):
        for c in range(1, min(max_col, ws.max_column + 1)):
            cell_value = ws.cell(row=r, column=c).value
            if cell_value and "实际回款" in str(cell_value).strip():
                return c - 1
    return None


def _load_json(path: Path) -> List[Dict]:
    """加载JSON文件，不存在返回空列表"""
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def _save_json(path: Path, data: Any):
    """保存JSON文件"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)



def _detect_summary_config(project_root: Path) -> Dict[str, Any]:
    """从 data/、template/ 自动识别路径、表头、实际回款列、分公司 sheet、项目分类"""
    config = {}
    data_dir = project_root / "data"
    template_dir = project_root / "template"
    
    # 回款表路径
    xlsx_list = sorted([p for p in data_dir.glob("*.xlsx") if not p.name.startswith("~$")])
    non_summary = [p for p in xlsx_list if "汇总" not in p.stem]
    if non_summary:
        config["table_path"] = str(non_summary[0].resolve().relative_to(project_root))
    elif xlsx_list:
        config["table_path"] = str(xlsx_list[0].resolve().relative_to(project_root))
    
    # 表头
    if config.get("table_path"):
        table_full = project_root / config["table_path"]
        try:
            wb = load_workbook(table_full, read_only=True, data_only=True)
            row1 = next(wb.active.iter_rows(min_row=1, max_row=1, values_only=True), None)
            wb.close()
            if row1 and any(row1):
                config["headers"] = [str(c).strip() if c else "" for c in row1]
        except (PermissionError, Exception) as e:
            print(f"警告: 无法读取表头 {table_full}: {str(e)}")
    
    # 模板路径
    template_xlsx = sorted([p for p in template_dir.glob("*.xlsx") if not p.name.startswith("~$")])
    if template_xlsx:
        config["template_path"] = str(template_xlsx[0].resolve().relative_to(project_root))
    
    # 从模板识别配置
    if config.get("template_path"):
        template_full = project_root / config["template_path"]
        try:
            wb = load_workbook(template_full, read_only=True, data_only=True)
            all_sheets = wb.sheetnames
            summary_sheets = [s for s in all_sheets if "汇总" in s]
            config["branch_sheet_names"] = [s for s in all_sheets if s not in summary_sheets]
            
            # 实际回款列
            summary_total = next((s for s in summary_sheets if "总" in s and "详细" not in s), summary_sheets[0] if summary_sheets else None)
            summary_detail = next((s for s in summary_sheets if "详细" in s), summary_sheets[1] if len(summary_sheets) > 1 else summary_sheets[0] if summary_sheets else None)
            
            if summary_total:
                col = _find_col_actual(wb[summary_total])
                if col is not None:
                    config["col_actual_total"] = col
            
            if summary_detail:
                col = _find_col_actual(wb[summary_detail])
                if col is not None:
                    config["col_actual_detail"] = col
            
            if config.get("branch_sheet_names"):
                col = _find_col_actual(wb[config["branch_sheet_names"][0]], max_row=2)
                if col is not None:
                    config["col_actual_branch"] = col
            
            # 项目分类
            skip = {"小计", "合计", "汇总", "项目分类", "负责人", "项目名称", "项目类型"}
            projects = []
            seen = set()
            for sn in summary_sheets:
                ws = wb[sn]
                for r in range(2, min(30, ws.max_row + 1)):
                    s = (ws.cell(row=r, column=2).value or "").strip()
                    if s and s not in skip and s not in seen:
                        seen.add(s)
                        projects.append(s)
                if projects:
                    break
            
            if projects:
                config["project_order"] = projects
            
            wb.close()
        except (PermissionError, Exception) as e:
            print(f"警告: 无法读取模板文件 {template_full}: {str(e)}")
    
    return config


def save_summary_config(config_path: Optional[str] = None, project_root: Optional[str] = None) -> Dict[str, Any]:
    """检测并保存 summary_config.json，返回 config"""
    root = (Path(project_root) if project_root else Path(__file__).resolve().parent).resolve()
    path = Path(config_path or root / SUMMARY_CONFIG_FILENAME)
    config = _detect_summary_config(root)
    _save_json(path, config)
    print(f"已自动识别并保存配置到 {path}")
    return config


def load_summary_config(config_path: Optional[str] = None, project_root: Optional[str] = None) -> Dict[str, Any]:
    """加载 summary_config.json；不存在则返回本次检测结果"""
    root = (Path(project_root) if project_root else Path(__file__).resolve().parent).resolve()
    path = Path(config_path or root / SUMMARY_CONFIG_FILENAME)
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return _detect_summary_config(root)


def extra_information(project_root: Optional[str] = None) -> Dict[str, Any]:
    """
    从模板 Excel 提取公司信息，按照5个步骤执行：
    1. 检测并保存配置
    2. 读取模板Excel文件
    3. 使用LLM提取公司信息
    4. 保存提取结果
    5. 返回结果
    """
    root = (Path(project_root) if project_root else Path(__file__).resolve().parent).resolve()
    template_dir = root / "template"
    
    # 步骤1: 检测并保存配置
    config = save_summary_config(project_root=str(root))
    
    # 步骤2: 读取模板Excel文件
    if not template_dir.exists():
        return {
            "success": False,
            "messages": [f"模板目录不存在: {template_dir}"]
        }
    
    table_text = read_xlsx(str(template_dir))
    if not table_text.strip():
        return {
            "success": False,
            "messages": [f"模板目录中没有找到Excel文件: {template_dir}"]
        }
    
    # 步骤3: 使用LLM提取公司信息
    content = ""
    try:
        llm = ChatTongyi(temperature=0)
        prompt = EXTRACT_PROMPT.format(table_text=table_text)
        response = llm.invoke(prompt)
        content = response.content.strip()
        
        # 提取JSON部分
        json_match = re.search(r'\[.*\]', content, re.DOTALL)
        if json_match:
            companies_data = json.loads(json_match.group())
        else:
            # 尝试直接解析整个内容
            companies_data = json.loads(content)
        
        # 步骤4: 保存提取结果
        output_path = root / TEMPLATE_COMPANIES_FILENAME
        _save_json(output_path, companies_data)
        
        # 步骤5: 返回结果
        return {
            "success": True,
            "messages": [
                f"配置已保存到: {root / SUMMARY_CONFIG_FILENAME}",
                f"提取了 {len(companies_data)} 条公司信息",
                f"结果已保存到: {output_path}"
            ],
            "config": config,
            "companies": companies_data,
            "count": len(companies_data)
        }
    
    except json.JSONDecodeError as e:
        return {
            "success": False,
            "messages": [
                f"JSON解析失败: {str(e)}",
                f"LLM返回内容: {content[:500] if content else '无内容'}"
            ]
        }
    except Exception as e:
        return {
            "success": False,
            "messages": [f"提取过程出错: {str(e)}"]
        }


if __name__ == "__main__":
    result = extra_information()
    print(result["messages"])
