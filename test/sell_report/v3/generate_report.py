"""
汇总报表生成脚本
功能：
1. 读取明细表数据（回款表和合同表）
2. 生成两个汇总表：
   - 预计汇总（总）：按公司名称和项目类型汇总
   - 预计汇总（详细）：按公司名称、负责人、项目类型汇总，合并相同公司名称
3. 统一小计行颜色
"""

import pandas as pd
from pathlib import Path
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import time
import pymysql

# ================= 配置区域 =================
PROJECT_ROOT = Path.cwd()
DETAIL_TABLE_PATH = PROJECT_ROOT / "data" / "明细表.xlsx"

# 数据库配置
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "qwer1234",
    "database": "sell_report",
    "charset": "utf8mb4"
}

# 列名映射：英文 -> 中文（用于表格显示）
COLUMN_MAPPING = {
    'responsible_person': '负责人',
    'company_name': '公司名称',
    'project_type': '项目类型',
    'project_name': '项目名称',
    'estimated_possible_payment_at_start': '月初预计可能回款',
    'estimated_confirmed_payment_at_start': '月初预计确定回款',
    'possible_payment': '可能回款',
    'confirmed_payment': '确定回款',
    'actual_payment': '实际回款',
    'payment_node_confirmed': '回款节点确定',
    'payment_node': '回款节点',
    'unpaid_amount': '未回款金额',
    'incomplete_reason': '未完成原因',
    'solution': '解决办法',
    'estimated_possible_at_start': '月初预计可能合同',
    'estimated_confirmed_at_start': '月初预计确定合同',
    'possible_contract': '可能合同',
    'confirmed_contract': '确定合同',
    'actual_contract': '实际合同',
    'completion_status': '完成情况',
    'contract_node': '合同节点',
    'incomplete_contract_amount': '未完成合同金额',
    'incomplete_contract_reason': '未完成合同原因',
    'date': '日期',
}

# 列配置（使用英文列名）
COL_CONFIG = {
    'pay': ['estimated_possible_payment_at_start', 'estimated_confirmed_payment_at_start', 
            'possible_payment', 'confirmed_payment', 'actual_payment'],
    'con': ['estimated_possible_at_start', 'estimated_confirmed_at_start', 
            'possible_contract', 'confirmed_contract', 'actual_contract']
}

def translate_columns_to_chinese(df):
    """将 DataFrame 的列名从英文转换为中文"""
    if df is None or df.empty:
        return df
    df = df.copy()
    df.columns = [COLUMN_MAPPING.get(col, col) for col in df.columns]
    return df

# 颜色配置 - 按照模板样式
SUBTOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")  # 淡黄色用于所有小计行
TYPE_SUMMARY_FILL = PatternFill("solid", fgColor="E7E6E6")  # 浅紫色用于分公司汇总行
GRAND_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")  # 淡黄色用于公司总计（与小计一致）
HEADER_ORANGE_FILL = PatternFill("solid", fgColor="FFE699")  # 浅橙色用于表头（月初预计可能/确定、可能/确定列）
HEADER_BLUE_FILL = PatternFill("solid", fgColor="D9E1F2")  # 浅蓝色用于表头（实际列）

# 标题样式
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True, size=11)


# ================= 数据读取函数 =================

def read_detail_table(year=None, month=None):
    """从数据库读取明细表，返回两个 DataFrame 格式的表格（列名为英文）"""
    try:
        # 连接数据库
        conn = pymysql.connect(**DB_CONFIG)
        
        # 构建 WHERE 条件
        where_conditions = ["company_name IS NOT NULL AND company_name != ''"]
        
        # 如果提供了年份和月份，添加日期过滤
        if year and month:
            where_conditions.append(f"YEAR(date) = {year} AND MONTH(date) = {month}")
        
        where_clause = " AND ".join(where_conditions)
        
        # 读取回款表数据，使用英文列名
        payment_query = f"""
            SELECT responsible_person, company_name, project_type, project_name, 
                   estimated_possible_payment_at_start, estimated_confirmed_payment_at_start, 
                   possible_payment, confirmed_payment, actual_payment, payment_node_confirmed,
                   payment_node, unpaid_amount, incomplete_reason, solution
            FROM payment_records
            WHERE {where_clause}
        """
        payment_df = pd.read_sql(payment_query, conn)
        
        # 读取合同表数据，使用英文列名
        contract_query = f"""
            SELECT responsible_person, company_name, project_type, project_name, 
                   estimated_possible_at_start, estimated_confirmed_at_start, 
                   possible_contract, confirmed_contract, actual_contract, 
                   completion_status
            FROM contract_records
            WHERE {where_clause}
        """
        contract_df = pd.read_sql(contract_query, conn)
        
        conn.close()
        
        filter_info = f"（{year}年{month}月）" if year and month else "（全部）"
        print(f"从数据库读取数据成功{filter_info}:")
        print(f"  回款表: {len(payment_df)} 条记录")
        print(f"  合同表: {len(contract_df)} 条记录")
        
        return payment_df, contract_df
        
    except Exception as e:
        print(f"❌ 从数据库读取数据失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame(), pd.DataFrame()

def get_summary_by_company_type_from_db(table_name, numeric_cols, year=None, month=None):
    """
    从数据库按公司名称和项目类型汇总，包含小计行（使用英文列名）
    :param table_name: 表名 ('payment_records' 或 'contract_records')
    :param numeric_cols: 需要汇总的数值列列表（英文列名）
    :param year: 年份（可选，用于过滤数据）
    :param month: 月份（可选，用于过滤数据）
    :return: 汇总后的DataFrame，包含标记列 '_row_type'（列名为英文）
    """
    try:
        conn = pymysql.connect(**DB_CONFIG)
        
        # 构建 WHERE 条件（使用英文列名）
        where_conditions = ["company_name IS NOT NULL AND company_name != ''"]
        
        # 如果有日期字段，添加日期过滤
        if year and month:
            where_conditions.append(f"YEAR(date) = {year} AND MONTH(date) = {month}")
        
        where_clause = " AND ".join(where_conditions)
        
        # 构建 SQL 查询，使用 UNION ALL 添加小计行
        numeric_cols_str = ", ".join([f"SUM(`{col}`) AS `{col}`" for col in numeric_cols])
        
        # 明细汇总查询（使用英文列名）
        detail_query = f"""
            SELECT 
                company_name,
                project_type,
                {numeric_cols_str},
                'DETAIL' AS _row_type
            FROM {table_name}
            WHERE {where_clause}
            GROUP BY company_name, project_type
        """
        
        # 小计行查询（按公司汇总，使用英文列名）
        subtotal_query = f"""
            SELECT 
                company_name,
                '小计' AS project_type,
                {numeric_cols_str},
                'SUBTOTAL' AS _row_type
            FROM {table_name}
            WHERE {where_clause}
            GROUP BY company_name
        """
        
        # 使用 UNION ALL 合并明细和小计，然后按公司名称和行类型排序
        final_query = f"""
            SELECT * FROM (
                {detail_query}
                UNION ALL
                {subtotal_query}
            ) AS combined
            ORDER BY company_name, 
                     CASE WHEN _row_type = 'DETAIL' THEN 0 ELSE 1 END,
                     project_type
        """
        
        result_df = pd.read_sql(final_query, conn)
        conn.close()
        
        return result_df
        
    except Exception as e:
        print(f"❌ 数据库汇总失败 ({table_name}): {str(e)}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def get_summary_by_person_from_db(table_name, numeric_cols, year=None, month=None):
    """
    从数据库按公司名称、负责人、项目类型汇总（使用英文列名）
    :param table_name: 表名 ('payment_records' 或 'contract_records')
    :param numeric_cols: 需要汇总的数值列列表（英文列名）
    :param year: 年份（可选，用于过滤数据）
    :param month: 月份（可选，用于过滤数据）
    :return: 汇总后的DataFrame（列名为英文）
    """
    try:
        conn = pymysql.connect(**DB_CONFIG)
        
        # 构建 WHERE 条件（使用英文列名）
        where_conditions = ["company_name IS NOT NULL AND company_name != ''"]
        
        # 如果有日期字段，添加日期过滤
        if year and month:
            where_conditions.append(f"YEAR(date) = {year} AND MONTH(date) = {month}")
        
        where_clause = " AND ".join(where_conditions)
        
        numeric_cols_str = ", ".join([f"SUM(`{col}`) AS `{col}`" for col in numeric_cols])
        
        query = f"""
            SELECT 
                company_name,
                responsible_person,
                project_type,
                {numeric_cols_str}
            FROM {table_name}
            WHERE {where_clause}
            GROUP BY company_name, responsible_person, project_type
            ORDER BY company_name, responsible_person, project_type
        """
        
        result_df = pd.read_sql(query, conn)
        conn.close()
        
        return result_df
        
    except Exception as e:
        print(f"❌ 数据库汇总失败 ({table_name}): {str(e)}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()


# ================= 第一部分汇总：按公司名称和项目类型汇总 =================

def generate_summary_by_company_type(df, numeric_cols, table_name=None, year=None, month=None):
    """
    第一部分汇总函数：按公司名称和项目类型汇总，并添加公司小计
    优先使用数据库汇总，如果 table_name 提供则从数据库查询
    :param df: 原始表 (pay_table 或 con_table)，如果从数据库查询则可为 None（列名为英文）
    :param numeric_cols: 需要累加的数值列清单（英文列名）
    :param table_name: 数据库表名 ('payment_records' 或 'contract_records')，如果提供则从数据库查询
    :param year: 年份（可选，用于过滤数据）
    :param month: 月份（可选，用于过滤数据）
    :return: 汇总后的DataFrame，包含标记列 '_row_type'（列名为中文）
    """
    # 如果提供了表名，从数据库查询（返回英文列名）
    if table_name:
        result = get_summary_by_company_type_from_db(table_name, numeric_cols, year, month)
        # 转换为中文列名
        return translate_columns_to_chinese(result)
    
    # 否则使用 pandas 处理（兼容旧逻辑，假设 df 列名已经是中文）
    if df is None or df.empty:
        return pd.DataFrame()
    
    # 1. 只提取必要的列，自动过滤掉负责人、项目名称等
    cols_to_keep = ['公司名称', '项目类型'] + numeric_cols
    # 过滤掉不在原表中的列，防止报错
    df_filtered = df[[c for c in cols_to_keep if c in df.columns]].copy()
    
    # 2. 按公司名称和项目类型进行初步汇总
    summary = df_filtered.groupby(['公司名称', '项目类型'])[numeric_cols].sum(numeric_only=True).reset_index()
    
    final_list = []
    # 3. 按照公司名称分组，循环插入小计行
    for company, group in summary.groupby('公司名称'):
        # 添加该公司的明细汇总数据
        group_copy = group.copy()
        group_copy['_row_type'] = 'DETAIL'
        final_list.append(group_copy)
        
        # 计算该公司的小计行
        subtotal = group[numeric_cols].sum().to_frame().T
        subtotal['公司名称'] = company  # 保持原始公司名称，以便合并单元格
        subtotal['项目类型'] = "小计"  # 小计行项目类型显示为"小计"
        subtotal['_row_type'] = 'SUBTOTAL'
        
        final_list.append(subtotal)
    
    # 4. 合并所有行
    result = pd.concat(final_list, ignore_index=True)
    return result


def generate_part1_summary(pay_table, con_table, year=None, month=None):
    """
    生成第一部分汇总：按公司名称和项目类型汇总
    使用数据库查询完成汇总
    :param year: 年份（可选，用于过滤数据）
    :param month: 月份（可选，用于过滤数据）
    :return: 合并后的汇总表（列名为中文）
    """
    # 1. 从数据库处理回款表汇总（返回中文列名）
    pay_cols = COL_CONFIG['pay']
    pay_summary_result = generate_summary_by_company_type(None, pay_cols, table_name='payment_records', year=year, month=month)
    
    # 2. 从数据库处理合同表汇总（返回中文列名）
    con_cols = COL_CONFIG['con']
    con_summary_result = generate_summary_by_company_type(None, con_cols, table_name='contract_records', year=year, month=month)
    
    # 3. 合并两表（使用中文列名）
    pay_cols_cn = [COLUMN_MAPPING.get(col, col) for col in pay_cols]
    con_cols_cn = [COLUMN_MAPPING.get(col, col) for col in con_cols]
    combined_result = pd.merge(
        pay_summary_result, 
        con_summary_result, 
        on=['公司名称', '项目类型'], 
        how='outer'
    )
    
    # 4. 清理数据：合并后，如果某一边没有数据会显示为 NaN，填充为 0
    combined_result = combined_result.fillna(0)
    
    # 5. 处理 _row_type 列：如果两边都有标记，保留其中一个
    if '_row_type_x' in combined_result.columns and '_row_type_y' in combined_result.columns:
        combined_result['_row_type'] = combined_result['_row_type_x'].fillna(combined_result['_row_type_y'])
        combined_result = combined_result.drop(columns=['_row_type_x', '_row_type_y'])
    elif '_row_type_x' in combined_result.columns:
        combined_result['_row_type'] = combined_result['_row_type_x']
        combined_result = combined_result.drop(columns=['_row_type_x'])
    elif '_row_type_y' in combined_result.columns:
        combined_result['_row_type'] = combined_result['_row_type_y']
        combined_result = combined_result.drop(columns=['_row_type_y'])
    
    return combined_result


# ================= 第二部分汇总：按公司名称、负责人、项目类型汇总 =================

def process_table_by_person(df, numeric_cols, table_name=None, year=None, month=None):
    """
    按公司名称、负责人、项目类型进行初步汇总
    优先使用数据库汇总，如果 table_name 提供则从数据库查询
    :param df: 原始表（列名为英文），如果从数据库查询则可为 None
    :param numeric_cols: 需要汇总的数值列列表（英文列名）
    :param table_name: 数据库表名，如果提供则从数据库查询
    :param year: 年份（可选，用于过滤数据）
    :param month: 月份（可选，用于过滤数据）
    :return: 汇总后的DataFrame（列名为中文）
    """
    # 如果提供了表名，从数据库查询（返回英文列名）
    if table_name:
        result = get_summary_by_person_from_db(table_name, numeric_cols, year, month)
        # 转换为中文列名
        return translate_columns_to_chinese(result)
    
    # 否则使用 pandas 处理（兼容旧逻辑，假设 df 列名已经是中文）
    if df is None or df.empty:
        return pd.DataFrame()
    
    keys = ['公司名称', '负责人', '项目类型']
    available_cols = [c for c in (keys + numeric_cols) if c in df.columns]
    df_filtered = df[available_cols].copy()
    return df_filtered.groupby(keys)[numeric_cols].sum(numeric_only=True).reset_index()


def generate_part2_summary(pay_table, con_table, year=None, month=None):
    """
    生成第二部分汇总：按公司名称、负责人、项目类型汇总
    使用数据库查询完成初步汇总
    :param year: 年份（可选，用于过滤数据）
    :param month: 月份（可选，用于过滤数据）
    :return: 汇总后的DataFrame，包含标记列 '_row_type'（列名为中文）
    """
    # 1. 定义列名（英文）
    pay_cols = COL_CONFIG['pay']
    con_cols = COL_CONFIG['con']
    
    # 转换为中文列名
    pay_cols_cn = [COLUMN_MAPPING.get(col, col) for col in pay_cols]
    con_cols_cn = [COLUMN_MAPPING.get(col, col) for col in con_cols]
    all_numeric_cols_cn = pay_cols_cn + con_cols_cn

    # 2. 从数据库获取初步汇总数据 (公司-负责人-项目)（返回中文列名）
    pay_summary = process_table_by_person(None, pay_cols, table_name='payment_records', year=year, month=month)
    con_summary = process_table_by_person(None, con_cols, table_name='contract_records', year=year, month=month)

    # 3. 拼接两表（使用中文列名）
    combined_result = pd.merge(
        pay_summary, con_summary, 
        on=['公司名称', '负责人', '项目类型'], 
        how='outer'
    ).fillna(0)

    # 4. 生成最终的三层汇总结构（这部分在 Python 中完成，因为需要插入汇总行）
    final_list = []

    # 按公司名称分组
    for company, group in combined_result.groupby('公司名称', sort=False):
        # --- A. 负责人明细层 ---
        # 确保排序是 [负责人 -> 项目类型]
        details = group.sort_values(['负责人', '项目类型']).copy()
        details['_row_type'] = 'DETAIL'
        final_list.append(details)
        
        # --- B. 公司各项目汇总层 (对应图片紫色部分) ---
        type_summary = group.groupby('项目类型')[all_numeric_cols_cn].sum().reset_index()
        type_summary['公司名称'] = company
        type_summary['负责人'] = company  # 负责人列显示公司名称（用于合并单元格）
        type_summary['_row_type'] = 'TYPE_SUMMARY'
        final_list.append(type_summary)
        
        # --- C. 公司最终小计 (对应图片黄色部分) ---
        subtotal = group[all_numeric_cols_cn].sum().to_frame().T
        subtotal['公司名称'] = company
        subtotal['负责人'] = "小计"
        subtotal['项目类型'] = "---"
        subtotal['_row_type'] = 'SUBTOTAL'
        final_list.append(subtotal)

    # 5. 合并所有层级
    full_df = pd.concat(final_list, ignore_index=True)
    return full_df


# ================= 第三部分汇总：按公司拆分，包含明细、类别汇总、公司总计 =================

def calculate_final_summaries(df, numeric_cols, remove_company_name=False):
    """
    逻辑：明细行展示 -> 底部按项目类型汇总 -> 底部公司总计
    :param df: 原始数据
    :param numeric_cols: 数值列
    :param remove_company_name: 是否删除公司名称列（用于分公司分表）
    """
    if df is None or df.empty:
        return pd.DataFrame()
    
    df = df.copy()
    df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
    
    # 1. 整理明细行 (按负责人、项目类型排序)
    details = df.sort_values(['负责人', '项目类型', '项目名称'] if '项目名称' in df.columns else ['负责人', '项目类型']).copy()
    details['_row_type'] = 'DETAIL'
    
    # 2. 计算【项目类别汇总】 (在所有明细最后展示)
    type_summary = df.groupby('项目类型')[numeric_cols].sum().reset_index()
    company_name = df['公司名称'].iloc[0] if '公司名称' in df.columns else ''
    if not remove_company_name:
        type_summary['公司名称'] = company_name
    # 负责人列显示公司名称（用于合并单元格），而不是"项目类别汇总"
    type_summary['负责人'] = company_name if company_name else '项目类别汇总'
    if '项目名称' in df.columns:
        type_summary['项目名称'] = '汇总'
    type_summary['_row_type'] = 'TYPE_SUMMARY'
    
    # 3. 计算【公司总计】
    grand_total = df[numeric_cols].sum().to_frame().T
    if not remove_company_name:
        grand_total['公司名称'] = df['公司名称'].iloc[0]
    grand_total['负责人'] = '公司总计'
    grand_total['项目类型'] = '---'
    if '项目名称' in df.columns:
        grand_total['项目名称'] = '---'
    grand_total['_row_type'] = 'GRAND_TOTAL'
    
    # 按照顺序拼接：明细 -> 类别汇总 -> 公司总计
    result = pd.concat([details, type_summary, grand_total], ignore_index=True)
    
    # 如果删除公司名称列，则从结果中移除
    if remove_company_name and '公司名称' in result.columns:
        result = result.drop(columns=['公司名称'])
    
    return result


# ================= Excel格式化函数 =================

def apply_excel_formatting(ws, df, title, start_row, merge_company=False, merge_person=True, year=None, month=None, add_borders=False):
    """
    格式化：合并标题、合并负责人、汇总行上色、合并公司名称、表头颜色
    :param ws: Excel工作表对象
    :param df: 要写入的DataFrame
    :param title: 表标题（如果为None则不添加标题）
    :param start_row: 起始行号（从1开始）
    :param merge_company: 是否合并相同公司名称（用于第二部分详细汇总）
    :param merge_person: 是否合并负责人列（默认True，第二个表设为False）
    :param year: 年份（用于表头颜色判断）
    :param month: 月份（用于表头颜色判断）
    :return: 下一张表的起始行号
    """
    if df.empty:
        return start_row
    
    # 排除标记列
    display_cols = [c for c in df.columns if c != '_row_type']
    max_col_idx = len(display_cols)
    
    # 确定数据起始行
    if title:
        data_header_row = start_row + 2
        data_start_row = data_header_row + 1
    else:
        # 没有标题时，直接从start_row+1开始（start_row+1是表头）
        data_start_row = start_row + 2
    
    # 1. 大标题处理（如果有标题）
    if title:
        ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=max_col_idx)
        cell = ws.cell(row=start_row + 1, column=1, value=title)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = TITLE_FONT
    
    # 2. 处理表头颜色
    if title:
        header_row = data_header_row
    else:
        header_row = start_row + 1
    
    # 根据列名判断表头颜色
    for col_idx, col_name in enumerate(display_cols, start=1):
        col_name_str = str(col_name)
        # 浅橙色：月初预计可能/确定、可能/确定列
        if any(keyword in col_name_str for keyword in ['月初预计可能', '月初预计确定', '可能回款', '确定回款', '可能合同', '确定合同']):
            ws.cell(header_row, col_idx).fill = HEADER_ORANGE_FILL
        # 浅蓝色：实际列
        elif '实际' in col_name_str:
            ws.cell(header_row, col_idx).fill = HEADER_BLUE_FILL
    
    # 找到公司名称列的索引（第一列）
    company_col_idx = 1
    
    # 2. 处理样式与合并逻辑
    m_start_company = None  # 用于合并公司名称
    m_start_person = None  # 用于合并负责人
    prev_person = None
    prev_row_type = None
    prev_company = None
    
    for i in range(len(df)):
        excel_row = data_start_row + i
        row_data = df.iloc[i]
        row_type = row_data.get('_row_type', 'DETAIL')
        company_name = row_data.get('公司名称', '')
        curr_person = row_data.get('负责人', '') if '负责人' in display_cols else ''
        
        # A. 上色逻辑 - 按照模板颜色
        if row_type == 'SUBTOTAL' or row_type == 'GRAND_TOTAL':
            # 小计和总计使用淡黄色
            for c in range(1, max_col_idx + 1):
                ws.cell(excel_row, c).fill = SUBTOTAL_FILL
        elif row_type == 'TYPE_SUMMARY':
            # 分公司汇总行使用浅紫色
            for c in range(1, max_col_idx + 1):
                ws.cell(excel_row, c).fill = TYPE_SUMMARY_FILL
        
        # B. 合并公司名称逻辑（如果启用）
        if merge_company:
            if i == 0:
                m_start_company = excel_row
                prev_company = company_name
            else:
                # 检查是否还是同一个公司
                if company_name == prev_company:
                    # 继续合并（同一个公司）
                    pass
                else:
                    # 公司名称变化，先合并之前的公司名称
                    if m_start_company and excel_row > m_start_company:
                        ws.merge_cells(start_row=m_start_company, start_column=company_col_idx, 
                                     end_row=excel_row - 1, end_column=company_col_idx)
                        ws.cell(m_start_company, company_col_idx).alignment = Alignment(
                            horizontal='center', vertical='center')
                    # 开始新的公司名称合并
                    m_start_company = excel_row
                    prev_company = company_name
            
            # 最后一行需要合并
            if i == len(df) - 1 and m_start_company:
                ws.merge_cells(start_row=m_start_company, start_column=company_col_idx, 
                             end_row=excel_row, end_column=company_col_idx)
                ws.cell(m_start_company, company_col_idx).alignment = Alignment(
                    horizontal='center', vertical='center')
        
        # C. 负责人合并逻辑 (第二列，如果有负责人列，且启用合并)
        if '负责人' in display_cols and merge_person:
            person_col_idx = display_cols.index('负责人') + 1
            
            # 初始化第一行
            if i == 0:
                m_start_person = excel_row
                prev_person = curr_person
                prev_row_type = row_type
                prev_company = company_name
            else:
                # 判断是否需要合并：
                # 1. 负责人内容相同
                # 2. 行类型相同（都是DETAIL或都是TYPE_SUMMARY）
                # 3. 公司名称相同（避免跨公司合并）
                should_merge = (curr_person == prev_person and 
                              row_type == prev_row_type and
                              company_name == prev_company and
                              row_type in ['DETAIL', 'TYPE_SUMMARY'])  # 只合并明细行和项目类别汇总行
                
                if not should_merge:
                    # 内容、类型或公司变化，先合并之前的单元格
                    if m_start_person and excel_row > m_start_person:
                        ws.merge_cells(start_row=m_start_person, start_column=person_col_idx, 
                                     end_row=excel_row - 1, end_column=person_col_idx)
                        ws.cell(m_start_person, person_col_idx).alignment = Alignment(
                            horizontal='center', vertical='center')
                    # 开始新的合并组
                    m_start_person = excel_row
                
                # 更新前一行信息
                prev_person = curr_person
                prev_row_type = row_type
                prev_company = company_name
            
            # 最后一行需要合并
            if i == len(df) - 1 and m_start_person:
                # 检查最后一行是否应该与前面的合并
                if (curr_person == prev_person and 
                    row_type == prev_row_type and 
                    company_name == prev_company and
                    row_type in ['DETAIL', 'TYPE_SUMMARY']):
                    ws.merge_cells(start_row=m_start_person, start_column=person_col_idx, 
                                 end_row=excel_row, end_column=person_col_idx)
                    ws.cell(m_start_person, person_col_idx).alignment = Alignment(
                        horizontal='center', vertical='center')
                elif excel_row > m_start_person:
                    # 最后一行单独，合并前面的
                    ws.merge_cells(start_row=m_start_person, start_column=person_col_idx, 
                                 end_row=excel_row - 1, end_column=person_col_idx)
                    ws.cell(m_start_person, person_col_idx).alignment = Alignment(
                        horizontal='center', vertical='center')

    # 3. 添加实线边框（如果启用）
    if add_borders:
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 确定边框范围：从表头行开始，到数据最后一行结束
        if title:
            border_start_row = data_header_row  # 表头行（标题行之后）
            border_end_row = data_start_row + len(df) - 1  # 数据最后一行
        else:
            border_start_row = start_row + 1  # 表头行
            border_end_row = data_start_row + len(df) - 1  # 数据最后一行
        
        # 为所有单元格添加边框（包括表头和数据行）
        for row_idx in range(border_start_row, border_end_row + 1):
            for col_idx in range(1, max_col_idx + 1):
                ws.cell(row_idx, col_idx).border = thin_border
    
    # 4. 隐藏标记列（如果存在）
    if '_row_type' in df.columns:
        col_idx = df.columns.get_loc('_row_type') + 1
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    
    return start_row + len(df) + (3 if title else 2)


def write_dataframe_to_excel(writer, df, sheet_name, start_row, title=None, merge_company=False, merge_person=True, year=None, month=None, add_borders=False):
    """
    将DataFrame写入Excel，并应用格式化
    :param writer: ExcelWriter对象
    :param df: 要写入的DataFrame
    :param sheet_name: 工作表名称
    :param start_row: 起始行号
    :param title: 表标题（None表示不添加标题）
    :param merge_company: 是否合并相同公司名称
    :param year: 年份（用于表头颜色判断）
    :param month: 月份（用于表头颜色判断）
    :param add_borders: 是否添加实线边框
    :return: 下一张表的起始行号
    """
    if df.empty:
        return start_row
    
    # 排除标记列
    display_df = df[[c for c in df.columns if c != '_row_type']].copy()
    
    # 确定写入起始行
    if title:
        write_start_row = start_row + 1
    else:
        write_start_row = start_row
    
    # 写入数据
    display_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=write_start_row)
    
    # 应用格式化
    ws = writer.book[sheet_name]
    return apply_excel_formatting(ws, df, title, start_row, merge_company, merge_person, year, month, add_borders)


# ================= 主流程函数 =================

def generate_complete_report(pay_table, con_table, year, month):
    """
    生成完整的汇总报表，包含三个部分：
    1. 预计汇总（总）：按公司名称和项目类型汇总
    2. 预计汇总（详细）：按公司名称、负责人、项目类型汇总，合并相同公司名称
    3. 按公司拆分：每个公司一个sheet，包含明细、类别汇总、公司总计
    :param pay_table: 回款表（列名为英文）
    :param con_table: 合同表（列名为英文）
    :param year: 年份
    :param month: 月份
    """
    # 如果表是英文列名，转换为中文列名用于后续处理
    if pay_table is not None and not pay_table.empty:
        pay_table_cn = translate_columns_to_chinese(pay_table.copy())
    else:
        pay_table_cn = pd.DataFrame()
    
    if con_table is not None and not con_table.empty:
        con_table_cn = translate_columns_to_chinese(con_table.copy())
    else:
        con_table_cn = pd.DataFrame()
    
    # 生成两个部分的汇总数据（函数内部会处理列名转换）
    part1_summary = generate_part1_summary(pay_table, con_table, year, month)
    part2_summary = generate_part2_summary(pay_table, con_table, year, month)
    
    # 按公司拆分数据用于第三部分（使用中文列名）
    pay_dict = {}
    if pay_table_cn is not None and not pay_table_cn.empty:
        pay_table_cn = pay_table_cn[pay_table_cn['公司名称'].notna() & (pay_table_cn['公司名称'].astype(str).str.strip() != "")]
        pay_dict = {n: pay_table_cn[pay_table_cn['公司名称'] == n].copy() for n in pay_table_cn['公司名称'].unique() if n and str(n).strip() != ""}
    
    con_dict = {}
    if con_table_cn is not None and not con_table_cn.empty:
        con_table_cn = con_table_cn[con_table_cn['公司名称'].notna() & (con_table_cn['公司名称'].astype(str).str.strip() != "")]
        con_dict = {n: con_table_cn[con_table_cn['公司名称'] == n].copy() for n in con_table_cn['公司名称'].unique() if n and str(n).strip() != ""}
    
    all_companies = sorted([c for c in set(pay_dict.keys()) | set(con_dict.keys()) if c and str(c).strip() != ""])
    
    # 生成输出文件名
    output_file = PROJECT_ROOT / "data" / f"{year}年{month}月份汇总表.xlsx"
    
    # 确保输出目录存在
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # 如果文件存在且被锁定，尝试删除或重命名
    if output_file.exists():
        try:
            # 尝试删除文件
            output_file.unlink()
            print(f"已删除旧文件: {output_file}")
        except PermissionError:
            # 如果无法删除（文件被打开），尝试重命名
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            backup_file = output_file.parent / f"{output_file.stem}_备份_{timestamp}{output_file.suffix}"
            try:
                output_file.rename(backup_file)
                print(f"文件被占用，已重命名为: {backup_file}")
            except Exception as e:
                print(f"错误: 无法删除或重命名文件 {output_file}")
                print(f"请关闭Excel文件后重试，或手动删除该文件")
                print(f"错误详情: {e}")
                return
        except Exception as e:
            print(f"处理旧文件时出错: {e}")
    
    # 创建Excel文件
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 创建第一个工作表：预计汇总（总）
            if not part1_summary.empty:
                title = f"{year}年{month}月份汇总表"
                write_dataframe_to_excel(
                    writer, 
                    part1_summary, 
                    "预计汇总（总）", 
                    0, 
                    title=title,
                    merge_company=True,  # 合并相同公司名称
                    year=year,
                    month=month,
                    add_borders=True  # 添加实线边框
                )
            
            # 创建第二个工作表：预计汇总（详细），没有标题，合并公司名称，合并负责人
            if not part2_summary.empty:
                write_dataframe_to_excel(
                    writer, 
                    part2_summary, 
                    "预计汇总（详细）", 
                    0, 
                    title=None,  # 没有标题
                    merge_company=True,  # 合并相同公司名称
                    merge_person=True,  # 合并相同负责人列
                    year=year,
                    month=month,
                    add_borders=True  # 添加实线边框
                )
            
            # 创建第三部分：按公司拆分，每个公司一个sheet
            for company in all_companies:
                # 过滤空的公司名称
                if not company or str(company).strip() == "":
                    continue
                
                sheet_name = str(company).replace("/", "_").replace("\\", "_")[:31]
                # 确保 sheet_name 不为空
                if not sheet_name or sheet_name.strip() == "":
                    sheet_name = f"公司_{len(all_companies)}"
                
                curr_row = 0
                
                # 处理回款块 - 表头改为"月份预计回款"，删除公司名称列，添加边框
                # 转换为中文列名
                pay_cols_cn = [COLUMN_MAPPING.get(col, col) for col in COL_CONFIG['pay']]
                df_p = calculate_final_summaries(pay_dict.get(company), pay_cols_cn, remove_company_name=True)
                if not df_p.empty:
                    curr_row = write_dataframe_to_excel(
                        writer, 
                        df_p, 
                        sheet_name, 
                        curr_row, 
                        title=f"{month}月份预计回款",
                        merge_company=False,
                        year=year,
                        month=month,
                        add_borders=True  # 添加实线边框
                    )
                
                # 处理合同块 - 表头改为"月份预计合同"，删除公司名称列，添加边框
                # 转换为中文列名
                con_cols_cn = [COLUMN_MAPPING.get(col, col) for col in COL_CONFIG['con']]
                df_c = calculate_final_summaries(con_dict.get(company), con_cols_cn, remove_company_name=True)
                if not df_c.empty:
                    write_dataframe_to_excel(
                        writer, 
                        df_c, 
                        sheet_name, 
                        curr_row, 
                        title=f"{month}月份预计合同",
                        merge_company=False,
                        year=year,
                        month=month,
                        add_borders=True  # 添加实线边框
                    )
        
        print(f"✅ 报表已生成！\n路径: {output_file}")
    except PermissionError as e:
        print(f"❌ 错误: 无法写入文件 {output_file}")
        print(f"请确保文件未被Excel或其他程序打开，然后重试")
        print(f"错误详情: {e}")
        return
    except Exception as e:
        print(f"❌ 生成报表时出错: {e}")
        import traceback
        traceback.print_exc()
        return


# ================= 主函数 =================

def generate_report_main(year=None, month=None, context=None):
    """
    主函数
    :param year: 年份，如2026
    :param month: 月份，如2
    """
    # 如果没有提供参数，从命令行获取或使用默认值
    if year is None or month is None:
        import sys
        if len(sys.argv) >= 3:
            try:
                year = int(sys.argv[1])
                month = int(sys.argv[2])
            except ValueError:
                print("错误: 年份和月份必须是整数")
                print("使用方法: python sum_table.py <年份> <月份>")
                print("示例: python sum_table.py 2026 2")
                return
        else:
            # 使用当前日期作为默认值
            from datetime import datetime
            now = datetime.now()
            year = now.year
            month = now.month
            print(f"未指定年份和月份，使用当前日期: {year}年{month}月")
    
    print(f"开始生成 {year}年{month}月 汇总报表...")
    print("开始读取明细表数据...")
    pay_table, con_table = read_detail_table(year, month)
    
    if pay_table is None or pay_table.empty:
        print("警告: 回款表数据为空")
    else:
        print(f"回款表数据: {len(pay_table)} 条记录")
    
    if con_table is None or con_table.empty:
        print("警告: 合同表数据为空")
    else:
        print(f"合同表数据: {len(con_table)} 条记录")
    
    print("\n开始生成汇总报表...")
    generate_complete_report(pay_table, con_table, year, month)
    print("\n完成！")


if __name__ == "__main__":
    generate_report_main(2026, 2,context="生成2026年2月份汇总报表")
